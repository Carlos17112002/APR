from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_protect
from django.db.models import Sum, Avg
from datetime import date
from decimal import Decimal
from django.core.exceptions import FieldError

from clientes.models import Cliente
from empresas.models import Empresa
from empresas.decorators import permiso_requerido

# ============================================================================
# VISTAS PARA GESTIÓN DE CLIENTES (sin autenticación)
# ============================================================================

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Max
from decimal import Decimal
from empresas.models import Empresa
from clientes.models import Cliente, Contrato

def obtener_proximo_numero_contrato(alias_db):
    """
    Obtiene el próximo número de contrato (incremental) basado en el máximo existente.
    Solo considera números que sean enteros (ignora textos no numéricos).
    Retorna el siguiente número como string.
    """
    contratos = Contrato.objects.using(alias_db).exclude(numero_contrato__isnull=True).exclude(numero_contrato='')
    max_num = 0
    for c in contratos:
        try:
            num = int(c.numero_contrato)
            if num > max_num:
                max_num = num
        except ValueError:
            continue
    return str(max_num + 1)

@permiso_requerido('crear_cliente')
def crear_cliente(request, alias):
    slug = alias
    alias_db = f'db_{slug}'
    empresa = get_object_or_404(Empresa, slug=slug)
    sectores = empresa.sectores()
    error = None

    if request.method == 'POST':
        # --- Campos de Cliente (básicos y adicionales) ---
        rut = request.POST.get('rut')
        nombre = request.POST.get('nombre')
        apellido_paterno = request.POST.get('apellido_paterno', '')
        apellido_materno = request.POST.get('apellido_materno', '')
        direccion = request.POST.get('direccion')
        medidor = request.POST.get('medidor')
        email = request.POST.get('email', '')
        telefono = request.POST.get('telefono', '')
        sector = request.POST.get('sector', '')
        latitude = request.POST.get('latitude')
        longitude = request.POST.get('longitude')

        # Sanitizar coordenadas
        if latitude:
            latitude = latitude.replace(',', '.')
            try:
                latitude = float(latitude)
            except (ValueError, TypeError):
                latitude = None
        if longitude:
            longitude = longitude.replace(',', '.')
            try:
                longitude = float(longitude)
            except (ValueError, TypeError):
                longitude = None

        fecha_nacimiento = request.POST.get('fecha_nacimiento') or None
        fecha_defuncion = request.POST.get('fecha_defuncion') or None
        sexo = request.POST.get('sexo', '')
        estado_civil = request.POST.get('estado_civil', '')
        profesion = request.POST.get('profesion', '')
        email_contacto = request.POST.get('email_contacto', '')
        contacto1 = request.POST.get('contacto1', '')
        contacto2 = request.POST.get('contacto2', '')
        fecha_incorporacion = request.POST.get('fecha_incorporacion') or None
        numero_libro = request.POST.get('numero_libro', '')

        # --- Campos del Contrato ---
        tipo_cliente = request.POST.get('tipo_cliente', '')
        tipo_servicio_ssr = request.POST.get('tipo_servicio_ssr', '')
        fecha_contrato = request.POST.get('fecha_contrato') or None
        comuna = request.POST.get('comuna', '')
        ciudad = request.POST.get('ciudad', '')
        direccion_arranque = request.POST.get('direccion_arranque', '')
        utm_norte = request.POST.get('utm_norte', '')
        utm_este = request.POST.get('utm_este', '')
        rol = request.POST.get('rol', '')
        socio = request.POST.get('socio') == 'on'
        servicio = request.POST.get('servicio', '')
        diametro = request.POST.get('diametro', '')
        marca_medidor = request.POST.get('marca_medidor', '')
        numero_medidor_contrato = request.POST.get('numero_medidor', '')
        ano_medidor = request.POST.get('ano_medidor') or None
        tipo_medidor = request.POST.get('tipo_medidor', '')
        sello_medidor = request.POST.get('sello_medidor', '')
        codigo_union_domiciliaria = request.POST.get('codigo_union_domiciliaria', '')
        email_recepcion_documento = request.POST.get('email_recepcion_documento', '')
        tarifa = request.POST.get('tarifa', '')
        tipo_documento = request.POST.get('tipo_documento', '')
        tipo_servicio = request.POST.get('tipo_servicio', '')

        # Número de contrato (si viene del formulario)
        numero_contrato_manual = request.POST.get('numero_contrato', '').strip()

        # Validar campos obligatorios del cliente
        if not all([rut, nombre, direccion, medidor]):
            error = 'Los campos RUT, Nombre, Dirección y Medidor son obligatorios.'
        elif Cliente.objects.using(alias_db).filter(rut=rut).exists():
            error = 'Ya existe un cliente con ese RUT.'
        else:
            try:
                # Crear cliente
                cliente = Cliente.objects.using(alias_db).create(
                    usuario_id=None,
                    empresa_slug=slug,
                    nombre=nombre,
                    apellido_paterno=apellido_paterno,
                    apellido_materno=apellido_materno,
                    rut=rut,
                    direccion=direccion,
                    telefono=telefono,
                    email=email,
                    medidor=medidor,
                    latitude=latitude,
                    longitude=longitude,
                    sector=sector,
                    fecha_nacimiento=fecha_nacimiento,
                    fecha_defuncion=fecha_defuncion,
                    sexo=sexo,
                    estado_civil=estado_civil,
                    profesion=profesion,
                    email_contacto=email_contacto,
                    contacto1=contacto1,
                    contacto2=contacto2,
                    fecha_incorporacion=fecha_incorporacion,
                    numero_libro=numero_libro,
                )

                # Determinar número de contrato
                if numero_contrato_manual:
                    numero_contrato = numero_contrato_manual
                    # Validar unicidad si se proporcionó manualmente
                    if Contrato.objects.using(alias_db).filter(numero_contrato=numero_contrato).exists():
                        error = f'El número de contrato {numero_contrato} ya está en uso.'
                        cliente.delete()  # Eliminar cliente creado para no dejar datos huérfanos
                        raise Exception(error)
                else:
                    numero_contrato = obtener_proximo_numero_contrato(alias_db)

                # Crear contrato asociado
                Contrato.objects.using(alias_db).create(
                    cliente=cliente,
                    numero_contrato=numero_contrato,
                    tipo_cliente=tipo_cliente,
                    tipo_servicio_ssr=tipo_servicio_ssr,
                    fecha_contrato=fecha_contrato,
                    comuna=comuna,
                    ciudad=ciudad,
                    sector_arranque=sector,
                    direccion_arranque=direccion_arranque,
                    utm_norte=utm_norte,
                    utm_este=utm_este,
                    rol=rol,
                    socio=socio,
                    servicio=servicio,
                    diametro=diametro,
                    marca_medidor=marca_medidor,
                    numero_medidor=numero_medidor_contrato,
                    ano_medidor=ano_medidor,
                    tipo_medidor=tipo_medidor,
                    sello_medidor=sello_medidor,
                    codigo_union_domiciliaria=codigo_union_domiciliaria,
                    email_recepcion_documento=email_recepcion_documento,
                    tarifa=tarifa,
                    tipo_documento=tipo_documento,
                    tipo_servicio=tipo_servicio,
                )

                messages.success(request, f'Cliente creado exitosamente. N° Contrato: {numero_contrato}')
                return redirect('listado_clientes', alias=slug)

            except Exception as e:
                if error is None:
                    error = f'Error al registrar el cliente: {str(e)}'

    return render(request, 'crear_cliente.html', {
        'empresa': empresa,
        'slug': slug,
        'sectores': sectores,
        'error': error,
    })


from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from empresas.models import Empresa
from clientes.models import Cliente, Contrato
from boletas.models import Boleta
from empresas.decorators import permiso_requerido

@permiso_requerido('clientes')
def listado_clientes(request, alias):
    slug = alias
    db_alias = f'db_{slug}'
    empresa = get_object_or_404(Empresa, slug=slug)

    sectores = empresa.sectores()
    clientes = Cliente.objects.using(db_alias).all()

    # Obtener parámetros de búsqueda (incluyendo número de contrato)
    sector = request.GET.get('sector')
    rut = request.GET.get('rut')
    nombre = request.GET.get('nombre')
    numero_contrato = request.GET.get('numero_contrato')  # NUEVO

    # Aplicar filtros
    if sector:
        clientes = clientes.filter(sector=sector)
    if rut:
        clientes = clientes.filter(rut__icontains=rut)
    if nombre:
        clientes = clientes.filter(nombre__icontains=nombre)
    if numero_contrato:
        # Filtrar por número de contrato a través de la relación OneToOne
        clientes = clientes.filter(contrato__numero_contrato__icontains=numero_contrato)

    clientes = clientes.order_by('nombre')

    # Obtener contratos relacionados en una sola consulta (evita N+1)
    contratos = Contrato.objects.using(db_alias).filter(cliente__in=clientes)
    contratos_dict = {c.cliente_id: c for c in contratos}

    # Período actual y boletas
    hoy = timezone.now()
    periodo_actual = f"{hoy.year}-{hoy.month:02d}"
    boletas_del_mes = Boleta.objects.using(db_alias).filter(
        empresa_slug=empresa.slug,
        periodo=periodo_actual
    ).values('cliente_id', 'id')
    clientes_con_boleta = {b['cliente_id']: b['id'] for b in boletas_del_mes}

    # Enriquecer clientes con datos adicionales
    for cliente in clientes:
        cliente.boleta_mes = cliente.id in clientes_con_boleta
        cliente.boleta_id = clientes_con_boleta.get(cliente.id)
        contrato = contratos_dict.get(cliente.id)
        cliente.numero_contrato = contrato.numero_contrato if contrato else ''

    clientes_con_email = clientes.exclude(email='').count()
    
    permisos_usuario = []
    if request.user.is_authenticated:
        try:
            from empresas.models import PerfilAdmin
            perfil = PerfilAdmin.objects.get(usuario=request.user)
            permisos_usuario = perfil.permisos.get(alias, [])
        except (PerfilAdmin.DoesNotExist, KeyError):
            pass

    context = {
        'empresa': empresa,
        'slug': slug,
        'clientes': clientes,
        'sectores': sectores,
        'clientes_con_email': clientes_con_email,
        'filtro_numero_contrato': numero_contrato,  # para mantener el valor en el campo de búsqueda
        'permisos_usuario' : permisos_usuario,
    }
    return render(request, 'listado_clientes.html', context)


from django.shortcuts import render, get_object_or_404
from django.db.models import Avg, Sum, Q
from django.db.models.functions import TruncMonth
from django.utils import timezone
from datetime import timedelta, datetime
import json

from empresas.models import Empresa
from clientes.models import (
    Cliente, Contrato, CambioMedidor,
    Subsidio, CargoPermanente, Cargo, Descuento,
    Convenio, OtroIngreso, CorteReposicion
)
from lecturas.models import LecturaMovil
from boletas.models import Boleta


def detalle_cliente(request, alias, cliente_id):
    db_alias = f'db_{alias}'
    empresa = get_object_or_404(Empresa, slug=alias)
    cliente = get_object_or_404(Cliente.objects.using(db_alias), id=cliente_id)

    # Contrato asociado
    try:
        contrato = cliente.contrato
    except Contrato.DoesNotExist:
        contrato = None

    # Última lectura con consumo válido
    ultima_lectura = LecturaMovil.objects.filter(
        empresa_slug=alias,
        cliente=cliente_id,
        consumo__isnull=False,
        consumo__gt=0
    ).order_by('-fecha_lectura').first()

    # Cambios de medidor
    cambios_medidor_qs = CambioMedidor.objects.using(db_alias).filter(
        cliente=cliente
    ).order_by('-fecha_registro')

    cambios_medidor = []
    for cm in cambios_medidor_qs:
        cambios_medidor.append({
            'periodo': cm.periodo or '-',
            'marca_retirado': cm.medidor_retirado_marca or '-',
            'numero_retirado': cm.medidor_retirado_numero,
            'ano_retirado': cm.medidor_retirado_anio or '-',
            'fecha_lectura_anterior': cm.fecha_lectura_anterior,
            'lectura_anterior': float(cm.lectura_anterior) if cm.lectura_anterior else 0,
            'lectura_retiro': float(cm.lectura_retiro),
            'consumo': float(cm.consumo_final) if cm.consumo_final else 0,
            'fecha_instalacion': cm.fecha_instalacion,
            'numero_instalado': cm.medidor_nuevo_numero,
            'lectura_inicial': float(cm.lectura_inicial),
            'fecha_registro': cm.fecha_registro,
            'usuario': cm.usuario or '-',
        })

    # Lecturas (para tabla y gráfico)
    lecturas = LecturaMovil.objects.filter(
        empresa_slug=alias,
        cliente=cliente_id
    ).order_by('-fecha_lectura')

    # ---------------------------------------------------------------
    # NUEVOS MODELOS
    # ---------------------------------------------------------------
    # Subsidios
    subsidios_qs = Subsidio.objects.using(db_alias).filter(cliente=cliente).order_by('-fecha_inicio')
    subsidios = list(subsidios_qs.values(
        'subsidio', 'municipalidad', 'numero_decreto', 'tramo',
        'rut_beneficiario', 'beneficiario', 'vivienda',
        'fecha_decreto', 'fecha_inicio', 'fecha_vencimiento', 'fecha_encuesta'
    ))

    # Cargos Permanentes
    cargos_permanentes_qs = CargoPermanente.objects.using(db_alias).filter(cliente=cliente).order_by('-fecha')
    cargos_permanentes = list(cargos_permanentes_qs.values(
        'fecha', 'codigo', 'descripcion', 'monto', 'facturable', 'periodo', 'usuario'
    ))

    # Cargos
    cargos_qs = Cargo.objects.using(db_alias).filter(cliente=cliente).order_by('-fecha')
    cargos = list(cargos_qs.values(
        'fecha', 'codigo', 'descripcion', 'monto', 'facturable', 'periodo', 'usuario'
    ))

    # Descuentos
    descuentos_qs = Descuento.objects.using(db_alias).filter(cliente=cliente).order_by('-fecha')
    descuentos = list(descuentos_qs.values(
        'fecha', 'codigo', 'descripcion', 'monto', 'facturable', 'periodo', 'usuario'
    ))

    # Convenios
    convenios_qs = Convenio.objects.using(db_alias).filter(cliente=cliente).order_by('-fecha_convenio')
    convenios = list(convenios_qs.values(
        'convenio', 'descripcion', 'fecha_convenio', 'total_cuotas',
        'monto', 'facturable', 'usuario'
    ))

    # Otros Ingresos
    otros_ingresos_qs = OtroIngreso.objects.using(db_alias).filter(cliente=cliente).order_by('-fecha_pago')
    otros_ingresos = list(otros_ingresos_qs.values(
        'concepto', 'documento', 'fecha_pago', 'monto', 'comprobante'
    ))

    # Corte y Reposición
    corte_qs = CorteReposicion.objects.using(db_alias).filter(cliente=cliente).order_by('-fecha_corte')
    historico_corte = []
    anulaciones_corte = []
    for c in corte_qs:
        item = {
            'fecha_corte': c.fecha_corte,
            'operador_corte': c.operador_corte,
            'lectura_corte': float(c.lectura_corte),
            'tipo_corte': c.tipo_corte,
            'fecha_reposicion': c.fecha_reposicion,
            'operador_reposicion': c.operador_reposicion,
            'anulado': c.anulado,
            'fecha_anulacion': c.fecha_anulacion,
            'usuario': c.usuario_anulacion or '',
        }
        if c.anulado:
            anulaciones_corte.append(item)
        else:
            historico_corte.append(item)

    # ---------------------------------------------------------------
    # Pagos (si existe)
    try:
        from pagos.models import Pago
        pagos = Pago.objects.using(db_alias).filter(cliente=cliente).order_by('-fecha')
    except ImportError:
        pagos = []

    # Estadísticas
    consumo_promedio = lecturas.filter(consumo__isnull=False).aggregate(Avg('consumo'))['consumo__avg'] or 0
    total_pagado = pagos.aggregate(Sum('monto'))['monto__sum'] or 0 if pagos else 0
    deuda_actual = 0  # Ajusta según tu lógica

    # Gráfico de consumo (últimos 6 meses)
    rango = request.GET.get('rango', '6m')
    hoy = timezone.now()

    if rango == '6m':
        fecha_inicio = hoy - timedelta(days=180)
    elif rango == '1y':
        fecha_inicio = hoy - timedelta(days=365)
    else:
        fecha_inicio = None

    lecturas_historico = LecturaMovil.objects.filter(
        empresa_slug=alias,
        cliente=cliente_id,
        consumo__isnull=False
    )
    if fecha_inicio:
        lecturas_historico = lecturas_historico.filter(fecha_lectura__gte=fecha_inicio)

    consumo_por_mes = lecturas_historico.annotate(
        mes=TruncMonth('fecha_lectura')
    ).values('mes').annotate(
        total_consumo=Sum('consumo')
    ).order_by('mes')

    meses_espanol = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                     'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    fechas_grafico = []
    consumos_grafico = []

    consumo_dict = {}
    for item in consumo_por_mes:
        if item['mes']:
            key = item['mes'].strftime('%Y-%m')
            consumo_dict[key] = float(item['total_consumo'] or 0)

    if rango == 'all':
        claves_ordenadas = sorted(consumo_dict.keys())
        fechas_grafico = [f"{meses_espanol[int(k[5:7])-1]} '{k[2:4]}" for k in claves_ordenadas]
        consumos_grafico = [consumo_dict[k] for k in claves_ordenadas]
    else:
        num_meses = 6 if rango == '6m' else 12
        for i in range(num_meses - 1, -1, -1):
            fecha = hoy - timedelta(days=30 * i)
            mes_num = fecha.month
            anio_num = fecha.year
            mes_nombre = meses_espanol[mes_num - 1]
            anio_corto = str(anio_num)[2:]
            fechas_grafico.append(f"{mes_nombre} '{anio_corto}")
            key = f"{anio_num}-{mes_num:02d}"
            consumos_grafico.append(consumo_dict.get(key, 0.0))

    # Lecturas para tabla
    lecturas_completas = []
    for lectura in lecturas:
        fecha_lectura_anterior = getattr(lectura, 'fecha_lectura_anterior', None)
        lecturas_completas.append({
            'id': lectura.id,
            'periodo': lectura.fecha_lectura.strftime('%m/%Y') if lectura.fecha_lectura else '',
            'fecha_lectura_anterior': fecha_lectura_anterior,
            'lectura_anterior': float(lectura.lectura_anterior) if lectura.lectura_anterior else 0,
            'fecha_lectura_actual': lectura.fecha_lectura,
            'lectura_actual': float(lectura.lectura_actual) if lectura.lectura_actual else 0,
            'consumo': float(lectura.consumo) if lectura.consumo else 0,
            'cambio_medidor': getattr(lectura, 'cambio_medidor', False),
            'termino_medio': getattr(lectura, 'termino_medio', ''),
            'saldo_promedio_anterior': getattr(lectura, 'saldo_promedio_anterior', ''),
            'consumo_facturado': float(getattr(lectura, 'consumo_facturado', lectura.consumo or 0)),
            'abono_proximo_periodo': getattr(lectura, 'abono_proximo_periodo', ''),
            'codigo_lectura': getattr(lectura, 'codigo_lectura', ''),
        })

    # Documentos (boletas)
    boletas = Boleta.objects.using(db_alias).filter(cliente=cliente).order_by('-fecha_emision')
    documentos = []
    for b in boletas:
        documentos.append({
            'periodo': b.periodo or (b.fecha_emision.strftime('%m/%Y') if b.fecha_emision else ''),
            'monto': float(b.total) if b.total else 0,
            'documento': f"Boleta #{b.id}",
            'fecha_emision': b.fecha_emision,
            'fecha_vencimiento': b.fecha_vencimiento,
            'fecha_pago': b.fecha_pago,
            'consumo': float(b.consumo) if b.consumo else 0,
            'url': '',
            'pago': b.estado == 'pagada',
            'traza': '',
            'usuario': '',
        })
    permisos_usuario = []
    if request.user.is_authenticated:
        try:
            from empresas.models import PerfilAdmin
            perfil = PerfilAdmin.objects.get(usuario=request.user)
            permisos_usuario = perfil.permisos.get(alias, [])
        except (PerfilAdmin.DoesNotExist, KeyError):
            pass

    context = {
        'empresa': empresa,
        'slug': alias,
        'cliente': cliente,
        'contrato': contrato,
        'ultima_lectura': ultima_lectura,
        'cambios_medidor': cambios_medidor,
        'lecturas': lecturas,
        'pagos': pagos,
        'consumo_promedio': consumo_promedio,
        'total_pagado': total_pagado,
        'deuda_actual': deuda_actual,
        'fechas_grafico': json.dumps(fechas_grafico),
        'consumos_grafico': json.dumps(consumos_grafico),
        'rango_seleccionado': rango,
        'documentos': documentos,

        # NUEVOS CONTEXTOS
        'subsidios': subsidios,
        'cargos_permanentes': cargos_permanentes,
        'cargos': cargos,
        'descuentos': descuentos,
        'convenios': convenios,
        'otros_ingresos': otros_ingresos,
        'historico_corte': historico_corte,
        'anulaciones_corte': anulaciones_corte,

        # Variables existentes que ya tenías
        'lecturas_completas': lecturas_completas,
        'permisos_usuario': permisos_usuario,
    }
    return render(request, 'clientes/detalle_cliente.html', context)


from django.db.models import Q, Max

@permiso_requerido('editar_cliente')
def editar_cliente(request, alias, cliente_id):
    db_alias = f'db_{alias}'
    empresa = get_object_or_404(Empresa, slug=alias)
    cliente = get_object_or_404(Cliente.objects.using(db_alias), id=cliente_id)

    try:
        contrato = cliente.contrato
    except Contrato.DoesNotExist:
        contrato = Contrato(cliente=cliente)

    sectores = empresa.sectores()
    error = None

    # --- Obtener números de contrato sugeridos (excluyendo el actual) ---
    existing_contratos = Contrato.objects.using(db_alias).exclude(cliente=cliente).exclude(
        Q(numero_contrato__isnull=True) | Q(numero_contrato='')
    ).values_list('numero_contrato', flat=True).distinct()
    
    # Calcular el siguiente número disponible (máximo + 1)
    max_num = Contrato.objects.using(db_alias).exclude(numero_contrato__isnull=True).exclude(numero_contrato='').aggregate(Max('numero_contrato'))['numero_contrato__max']
    if max_num:
        try:
            next_num = int(max_num) + 1
            suggested_numbers = list(existing_contratos) + [str(next_num)]
        except ValueError:
            suggested_numbers = list(existing_contratos)
    else:
        suggested_numbers = list(existing_contratos) + ['1']

    if request.method == 'POST':
        # --- Campos de Cliente ---
        rut = request.POST.get('rut')
        nombre = request.POST.get('nombre')
        apellido_paterno = request.POST.get('apellido_paterno', '')
        apellido_materno = request.POST.get('apellido_materno', '')
        direccion = request.POST.get('direccion')
        medidor = request.POST.get('medidor')
        email = request.POST.get('email', '')
        telefono = request.POST.get('telefono', '')
        sector = request.POST.get('sector', '')
        latitude = request.POST.get('latitude')
        longitude = request.POST.get('longitude')

        # Sanitizar coordenadas
        if latitude:
            latitude = latitude.replace(',', '.')
            try:
                latitude = float(latitude)
            except (ValueError, TypeError):
                latitude = None
        if longitude:
            longitude = longitude.replace(',', '.')
            try:
                longitude = float(longitude)
            except (ValueError, TypeError):
                longitude = None

        fecha_nacimiento = request.POST.get('fecha_nacimiento') or None
        fecha_defuncion = request.POST.get('fecha_defuncion') or None
        sexo = request.POST.get('sexo', '')
        estado_civil = request.POST.get('estado_civil', '')
        profesion = request.POST.get('profesion', '')
        email_contacto = request.POST.get('email_contacto', '')
        contacto1 = request.POST.get('contacto1', '')
        contacto2 = request.POST.get('contacto2', '')
        fecha_incorporacion = request.POST.get('fecha_incorporacion') or None
        numero_libro = request.POST.get('numero_libro', '')

        # --- Campos del contrato ---
        tipo_cliente = request.POST.get('tipo_cliente', '')
        tipo_servicio_ssr = request.POST.get('tipo_servicio_ssr', '')
        fecha_contrato = request.POST.get('fecha_contrato') or None
        comuna = request.POST.get('comuna', '')
        ciudad = request.POST.get('ciudad', '')
        direccion_arranque = request.POST.get('direccion_arranque', '')
        utm_norte = request.POST.get('utm_norte', '')
        utm_este = request.POST.get('utm_este', '')
        rol = request.POST.get('rol', '')
        socio = request.POST.get('socio') == 'on'
        servicio = request.POST.get('servicio', '')
        diametro = request.POST.get('diametro', '')
        marca_medidor = request.POST.get('marca_medidor', '')
        numero_medidor_contrato = request.POST.get('numero_medidor', '')
        ano_medidor = request.POST.get('ano_medidor') or None
        tipo_medidor = request.POST.get('tipo_medidor', '')
        sello_medidor = request.POST.get('sello_medidor', '')
        codigo_union_domiciliaria = request.POST.get('codigo_union_domiciliaria', '')
        email_recepcion_documento = request.POST.get('email_recepcion_documento', '')
        tarifa = request.POST.get('tarifa', '')
        tipo_documento = request.POST.get('tipo_documento', '')
        tipo_servicio = request.POST.get('tipo_servicio', '')

        # --- Número de contrato (con validación de unicidad) ---
        numero_contrato = request.POST.get('numero_contrato', '').strip()

        # Validar campos obligatorios del cliente
        if not all([rut, nombre, direccion, medidor]):
            error = 'Los campos RUT, Nombre, Dirección y Medidor son obligatorios.'
        else:
            try:
                # Actualizar cliente
                cliente.rut = rut
                cliente.nombre = nombre
                cliente.apellido_paterno = apellido_paterno
                cliente.apellido_materno = apellido_materno
                cliente.direccion = direccion
                cliente.medidor = medidor
                cliente.email = email
                cliente.telefono = telefono
                cliente.sector = sector
                cliente.latitude = latitude
                cliente.longitude = longitude
                cliente.fecha_nacimiento = fecha_nacimiento
                cliente.fecha_defuncion = fecha_defuncion
                cliente.sexo = sexo
                cliente.estado_civil = estado_civil
                cliente.profesion = profesion
                cliente.email_contacto = email_contacto
                cliente.contacto1 = contacto1
                cliente.contacto2 = contacto2
                cliente.fecha_incorporacion = fecha_incorporacion
                cliente.numero_libro = numero_libro
                cliente.save(using=db_alias)

                # Actualizar contrato (validar número si se cambió)
                if numero_contrato and contrato.numero_contrato != numero_contrato:
                    if Contrato.objects.using(db_alias).filter(numero_contrato=numero_contrato).exists():
                        error = f'El número de contrato "{numero_contrato}" ya está en uso.'
                    else:
                        contrato.numero_contrato = numero_contrato
                elif not numero_contrato:
                    contrato.numero_contrato = None  # o puedes mantener el anterior si prefieres
                # Actualizar demás campos del contrato
                contrato.tipo_cliente = tipo_cliente
                contrato.tipo_servicio_ssr = tipo_servicio_ssr
                contrato.fecha_contrato = fecha_contrato
                contrato.comuna = comuna
                contrato.ciudad = ciudad
                contrato.sector_arranque = sector
                contrato.direccion_arranque = direccion_arranque
                contrato.utm_norte = utm_norte
                contrato.utm_este = utm_este
                contrato.rol = rol
                contrato.socio = socio
                contrato.servicio = servicio
                contrato.diametro = diametro
                contrato.marca_medidor = marca_medidor
                contrato.numero_medidor = numero_medidor_contrato
                contrato.ano_medidor = ano_medidor
                contrato.tipo_medidor = tipo_medidor
                contrato.sello_medidor = sello_medidor
                contrato.codigo_union_domiciliaria = codigo_union_domiciliaria
                contrato.email_recepcion_documento = email_recepcion_documento
                contrato.tarifa = tarifa
                contrato.tipo_documento = tipo_documento
                contrato.tipo_servicio = tipo_servicio

                if not error:
                    contrato.save(using=db_alias)
                    messages.success(request, f'Cliente {cliente.nombre} actualizado correctamente.')
                    return redirect('detalle_cliente', alias=alias, cliente_id=cliente.id)

            except Exception as e:
                error = f'Error al actualizar el cliente: {str(e)}'

    context = {
        'empresa': empresa,
        'slug': alias,
        'cliente': cliente,
        'contrato': contrato,
        'sectores': sectores,
        'error': error,
        'suggested_contrato_numbers': suggested_numbers,  # para el datalist
    }
    return render(request, 'clientes/editar_cliente.html', context)


def historial_cliente(request, alias, cliente_id):
    """
    Muestra el historial completo del cliente: todas las lecturas,
    pagos, contratos y avisos (avisos generales de la empresa).
    """
    db_alias = f'db_{alias}'
    empresa = get_object_or_404(Empresa, slug=alias)          # desde base default
    cliente = get_object_or_404(Cliente.objects.using(db_alias), id=cliente_id)

    # Importaciones condicionales con manejo de errores
    lecturas_completas = []
    pagos_completos = []
    contratos_completos = []
    avisos = []

    try:
        from lecturas.models import Lectura
        lecturas_completas = Lectura.objects.using(db_alias).filter(cliente=cliente).order_by('-fecha')
    except (ImportError, FieldError, Exception) as e:
        # Si no existe el modelo o el campo, dejar lista vacía
        logger.warning(f"Error al obtener lecturas: {e}")

    try:
        from pagos.models import Pago
        pagos_completos = Pago.objects.using(db_alias).filter(cliente=cliente).order_by('-fecha')
    except (ImportError, FieldError, Exception) as e:
        logger.warning(f"Error al obtener pagos: {e}")

    try:
        from contratos.models import Contrato
        contratos_completos = Contrato.objects.using(db_alias).filter(cliente=cliente).order_by('-fecha_inicio')
    except (ImportError, FieldError, Exception) as e:
        logger.warning(f"Error al obtener contratos: {e}")

    try:
        from avisos.models import Aviso
        # Suponiendo que Aviso tiene un campo 'empresa' (ForeignKey a Empresa)
        # y queremos mostrar avisos de la empresa (no específicos del cliente)
        avisos = Aviso.objects.using(db_alias).filter(empresa=empresa).order_by('-fecha_creacion')
    except (ImportError, FieldError, Exception) as e:
        logger.warning(f"Error al obtener avisos: {e}")
        # Si el campo es otro, intentar con 'empresa_id' o similar
        try:
            avisos = Aviso.objects.using(db_alias).filter(empresa_id=empresa.id).order_by('-fecha_creacion')
        except:
            avisos = []

    # Estadísticas (calcular solo si hay datos)
    consumo_total = 0
    pago_total = 0
    consumo_promedio = 0
    if lecturas_completas:
        consumo_total = lecturas_completas.aggregate(Sum('consumo'))['consumo__sum'] or 0
        consumo_promedio = lecturas_completas.aggregate(Avg('consumo'))['consumo__avg'] or 0
    if pagos_completos:
        pago_total = pagos_completos.aggregate(Sum('monto'))['monto__sum'] or 0

    total_registros = len(lecturas_completas) + len(pagos_completos) + len(contratos_completos) + len(avisos)

    context = {
        'empresa': empresa,
        'slug': alias,
        'cliente': cliente,
        'lecturas': lecturas_completas,
        'pagos': pagos_completos,
        'contratos': contratos_completos,
        'avisos': avisos,
        'consumo_total': consumo_total,
        'pago_total': pago_total,
        'consumo_promedio': consumo_promedio,
        'total_registros': total_registros,
    }

    return render(request, 'clientes/historial_cliente.html', context)


@require_POST
@csrf_protect
def eliminar_cliente(request, alias, cliente_id):
    """
    Elimina un cliente de la base de datos de la empresa.
    Soporta peticiones AJAX y normales.
    """
    db_alias = f'db_{alias}'
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    try:
        cliente = get_object_or_404(Cliente.objects.using(db_alias), id=cliente_id)
        cliente.delete()

        if is_ajax:
            return JsonResponse({'success': True, 'message': 'Cliente eliminado exitosamente.'})
        else:
            messages.success(request, f'Cliente {cliente.nombre} eliminado.')
            return redirect('listado_clientes', alias=alias)

    except Exception as e:
        if is_ajax:
            return JsonResponse({'success': False, 'error': str(e)})
        else:
            messages.error(request, f'Error al eliminar: {str(e)}')
            return redirect('listado_clientes', alias=alias)


def clientes_por_alias(request, alias):
    """
    API que devuelve los clientes con coordenadas (para mapas).
    """
    try:
        clientes = Cliente.objects.using(f'db_{alias}').exclude(latitude=None).exclude(longitude=None)
        data = list(clientes.values('id', 'nombre', 'direccion', 'latitude', 'longitude'))
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from boletas.models import Boleta
from empresas.models import Empresa

def ver_boleta_cliente(request, alias, cliente_id):
    # Obtener empresa desde la base por defecto (Empresa está en BD principal)
    empresa = get_object_or_404(Empresa, slug=alias)
    
    # Determinar el alias de la base de datos de la empresa
    db_alias = f'db_{alias}'
    
    # Obtener el período actual (YYYY-MM)
    hoy = timezone.now()
    periodo_actual = f"{hoy.year}-{hoy.month:02d}"
    
    # Buscar la boleta en la base de datos de la empresa
    boleta = get_object_or_404(
        Boleta.objects.using(db_alias),
        empresa_slug=empresa.slug,
        cliente_id=cliente_id,
        periodo=periodo_actual
    )
    
    return render(request, 'boletas/detalle_boleta.html', {
        'boleta': boleta,
        'empresa': empresa,
        'slug': alias,  # opcional, para usar en templates
    })

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from decimal import Decimal
from boletas.models import Boleta
from clientes.models import Cliente
from lecturas.models import LecturaMovil
from empresas.models import Empresa
from boletas.helpers import generar_boletas_por_alias  # si quieres usar la función masiva

def generar_boleta_individual(request, empresa_slug, cliente_id, lectura_id=None):
    """
    Genera una boleta para un cliente específico.
    Si se proporciona lectura_id, la asocia; si no, calcula con la última lectura.
    """
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    
    # Obtener cliente desde la base de datos de la empresa
    alias_db = f'db_{empresa_slug}'
    cliente = Cliente.objects.using(alias_db).get(id=cliente_id)
    
    # Si se proporciona lectura_id, verificar que exista
    lectura = None
    if lectura_id:
        lectura = LecturaMovil.objects.using(alias_db).get(id=lectura_id, cliente_id=cliente_id)
    
    # Verificar si ya existe una boleta para este período (mes actual)
    hoy = timezone.now()
    mes_actual = hoy.month
    año_actual = hoy.year
    periodo_str = f"{año_actual}-{mes_actual:02d}"
    
    # Buscar boleta existente (usando la base de datos default donde están las boletas)
    boleta_existente = Boleta.objects.filter(
        empresa_slug=empresa_slug,
        cliente_id=cliente_id,
        periodo=periodo_str
    ).first()
    
    if boleta_existente:
        messages.warning(request, f"Ya existe una boleta para {cliente.nombre} en el período {periodo_str}.")
        return redirect('ver_boleta', boleta_id=boleta_existente.id)
    
    # Calcular consumo
    # Si no hay lectura, usar la última lectura registrada
    if not lectura:
        # Obtener la última lectura (podrías tener un método en el modelo)
        lectura = LecturaMovil.objects.using(alias_db).filter(
            cliente_id=cliente_id
        ).order_by('-fecha_lectura').first()
        if not lectura:
            messages.error(request, "No hay lecturas para este cliente.")
            return redirect('listado_clientes', alias=empresa_slug)
    
    # Obtener lectura anterior (la penúltima)
    lectura_anterior = LecturaMovil.objects.using(alias_db).filter(
        cliente_id=cliente_id
    ).exclude(id=lectura.id).order_by('-fecha_lectura').first()
    
    valor_anterior = lectura_anterior.valor if lectura_anterior else 0
    valor_actual = lectura.valor
    consumo = valor_actual - valor_anterior
    if consumo < 0:
        consumo = 0  # o manejar reinicio de medidor
    
    # Calcular montos (esto debería estar en una función de negocio)
    tarifa_m3 = Decimal('500')  # ejemplo, debería venir de configuración
    cargo_fijo = Decimal('2000')
    monto_consumo = consumo * tarifa_m3
    total = monto_consumo + cargo_fijo
    
    # Crear boleta
    boleta = Boleta.objects.create(
        cliente_id=cliente_id,
        empresa_slug=empresa_slug,
        lectura_id=lectura.id,
        periodo=periodo_str,
        fecha_vencimiento=hoy.replace(day=10) + timezone.timedelta(days=30),  # ejemplo
        lectura_anterior=valor_anterior,
        lectura_actual=valor_actual,
        consumo=consumo,
        monto_consumo=monto_consumo,
        cargo_fijo=cargo_fijo,
        otros_cargos=0,
        total=total,
        estado='generada',
        codigo_barras='',  # generar si es necesario
    )
    
    messages.success(request, f"Boleta generada para {cliente.nombre} por ${total}.")
    return redirect('ver_boleta', boleta_id=boleta.id)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import date

def exportarclientes_excel(request, alias):
    """
    Exporta la lista de clientes a Excel con formato profesional.
    Respeta los filtros GET: sector, rut, nombre.
    """
    db_alias = f'db_{alias}'
    empresa = get_object_or_404(Empresa, slug=alias)

    # Obtener clientes con los mismos filtros que en el listado
    clientes = Cliente.objects.using(db_alias).all()

    sector = request.GET.get('sector')
    rut = request.GET.get('rut')
    nombre = request.GET.get('nombre')
    if sector:
        clientes = clientes.filter(sector=sector)
    if rut:
        clientes = clientes.filter(rut__icontains=rut)
    if nombre:
        clientes = clientes.filter(nombre__icontains=nombre)

    clientes = clientes.order_by('nombre')

    # Obtener contratos relacionados en una sola consulta (para número de contrato)
    contratos = Contrato.objects.using(db_alias).filter(cliente__in=clientes)
    contratos_dict = {c.cliente_id: c for c in contratos}

    # Crear libro y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    # Estilos (igual que en otros informes)
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Calibri', size=10)
    cell_alignment = Alignment(horizontal='left', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Encabezados (incluyendo N° Contrato)
    headers = [
        'ID', 'RUT', 'Nombre', 'Apellido Paterno', 'Apellido Materno',
        'Email', 'Teléfono', 'Dirección', 'Sector', 'N° Medidor',
        'N° Contrato', 'Latitud', 'Longitud'
    ]

    # Escribir encabezados
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Llenar datos
    for row_idx, cliente in enumerate(clientes, start=2):
        contrato = contratos_dict.get(cliente.id)
        row_data = [
            cliente.id,
            cliente.rut or '',
            cliente.nombre or '',
            cliente.apellido_paterno or '',
            cliente.apellido_materno or '',
            cliente.email or '',
            cliente.telefono or '',
            cliente.direccion or '',
            cliente.sector or '',
            cliente.medidor or '',
            contrato.numero_contrato if contrato else '',
            cliente.latitude or '',
            cliente.longitude or '',
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

    # Ajustar ancho de columnas
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row, col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Congelar primera fila (opcional)
    ws.freeze_panes = 'A2'

    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="clientes_{alias}_{date.today()}.xlsx"'
    wb.save(response)
    return response


import csv
import io
from datetime import datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from empresas.models import Empresa
from clientes.models import Cliente, Contrato

def importar_clientes(request, alias):
    db_alias = f'db_{alias}'
    empresa = get_object_or_404(Empresa, slug=alias)

    if request.method == 'POST':
        csv_file = request.FILES.get('csv_file')
        if not csv_file:
            messages.error(request, 'Debe seleccionar un archivo CSV.')
            return redirect('importar_clientes', alias=alias)

        if not csv_file.name.endswith('.csv'):
            messages.error(request, 'El archivo debe ser CSV.')
            return redirect('importar_clientes', alias=alias)

        try:
            # Leer archivo con UTF-8 con BOM
            data = csv_file.read().decode('utf-8-sig')
            io_string = io.StringIO(data)
            reader = csv.DictReader(io_string, delimiter=';')

            # Limpiar nombres de columnas
            reader.fieldnames = [name.strip() for name in reader.fieldnames]

            creados = 0
            actualizados = 0
            errores = []

            mapping_cliente = {
                'rut': 'rut',
                'nombre': 'nombre',
                'apellido_paterno': 'apellido_paterno',
                'apellido_materno': 'apellido_materno',
                'email': 'email',
                'telefono': 'telefono',
                'direccion': 'direccion',
                'medidor': 'medidor',
                'sector': 'sector',
                'latitude': 'latitude',
                'longitude': 'longitude',
                'fecha_nacimiento': 'fecha_nacimiento',
                'fecha_defuncion': 'fecha_defuncion',
                'sexo': 'sexo',
                'estado_civil': 'estado_civil',
                'profesion': 'profesion',
                'email_contacto': 'email_contacto',
                'contacto1': 'contacto1',
                'contacto2': 'contacto2',
                'fecha_incorporacion': 'fecha_incorporacion',
                'numero_libro': 'numero_libro',
            }

            mapping_contrato = {
                'tipo_cliente': 'tipo_cliente',
                'tipo_servicio_ssr': 'tipo_servicio_ssr',
                'fecha_contrato': 'fecha_contrato',
                'comuna': 'comuna',
                'ciudad': 'ciudad',
                'sector_arranque': 'sector_arranque',
                'direccion_arranque': 'direccion_arranque',
                'utm_norte': 'utm_norte',
                'utm_este': 'utm_este',
                'rol': 'rol',
                'socio': 'socio',
                'servicio': 'servicio',
                'diametro': 'diametro',
                'marca_medidor': 'marca_medidor',
                'numero_medidor': 'numero_medidor',
                'ano_medidor': 'ano_medidor',
                'tipo_medidor': 'tipo_medidor',
                'sello_medidor': 'sello_medidor',
                'codigo_union_domiciliaria': 'codigo_union_domiciliaria',
                'email_recepcion_documento': 'email_recepcion_documento',
                'tarifa': 'tarifa',
                'tipo_documento': 'tipo_documento',
                'tipo_servicio': 'tipo_servicio',
            }

            # Función para limpiar valores (quitar comillas)
            def limpiar_valor(val):
                if not val:
                    return ''
                val = val.strip()
                if val.startswith('"') and val.endswith('"'):
                    val = val[1:-1]
                return val

            # Función para parsear fechas
            def parse_fecha(val):
                if not val:
                    return None
                val = val.strip()
                # Quitar comillas
                if val.startswith('"') and val.endswith('"'):
                    val = val[1:-1]
                # Reemplazar guiones por barras para unificar
                val = val.replace('-', '/')
                # Intentar parsear
                try:
                    return datetime.strptime(val, '%d/%m/%Y').date()
                except ValueError:
                    try:
                        return datetime.strptime(val, '%Y/%m/%d').date()
                    except ValueError:
                        return None

            for fila_num, row in enumerate(reader, start=2):
                # Limpiar todos los valores
                row = {k: limpiar_valor(v) for k, v in row.items()}

                rut = row.get('rut', '').strip()
                if not rut:
                    errores.append(f"Fila {fila_num}: RUT vacío")
                    continue

                nombre = row.get('nombre', '').strip()
                direccion = row.get('direccion', '').strip()
                medidor = row.get('medidor', '').strip()

                if not (nombre and direccion and medidor):
                    errores.append(f"Fila {fila_num}: faltan campos obligatorios (nombre, direccion o medidor)")
                    continue

                # Datos del cliente
                cliente_data = {}
                for csv_key, model_field in mapping_cliente.items():
                    if csv_key in row and row[csv_key]:
                        value = row[csv_key]
                        if model_field in ['latitude', 'longitude']:
                            try:
                                value = float(value.replace(',', '.'))
                            except:
                                value = None
                        elif model_field in ['fecha_nacimiento', 'fecha_defuncion', 'fecha_incorporacion']:
                            value = parse_fecha(value)
                        cliente_data[model_field] = value

                cliente_data['empresa_slug'] = alias

                cliente, created = Cliente.objects.using(db_alias).update_or_create(
                    rut=rut,
                    defaults=cliente_data
                )

                if created:
                    creados += 1
                else:
                    actualizados += 1

                # Contrato
                contrato, _ = Contrato.objects.using(db_alias).get_or_create(cliente=cliente)
                contrato_data = {}
                for csv_key, model_field in mapping_contrato.items():
                    if csv_key in row and row[csv_key]:
                        value = row[csv_key]
                        if model_field == 'socio':
                            value = value.lower() in ('si', 'sí', 'true', '1', 'yes')
                        elif model_field == 'fecha_contrato':
                            value = parse_fecha(value)
                        contrato_data[model_field] = value

                if contrato_data:
                    for field, val in contrato_data.items():
                        setattr(contrato, field, val)
                    contrato.save(using=db_alias)

            messages.success(request, f'Importación completada: {creados} creados, {actualizados} actualizados.')
            if errores:
                messages.warning(request, f'Se encontraron {len(errores)} errores. Los primeros: {", ".join(errores[:3])}')

            return redirect('listado_clientes', alias=alias)

        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
            return redirect('importar_clientes', alias=alias)

    return render(request, 'importar_clientes.html', {'empresa': empresa, 'slug': alias})

from django.db import connections
from django.contrib.auth.decorators import login_required
import json

@login_required
def mapa_clientes(request, alias):
    """
    Muestra un mapa con la ubicación de todos los clientes de la empresa.
    """
    db_alias = f'db_{alias}'
    empresa = get_object_or_404(Empresa, slug=alias)
    
    # Obtener clientes con coordenadas válidas
    clientes = Cliente.objects.using(db_alias).filter(
        latitude__isnull=False,
        longitude__isnull=False
    ).exclude(latitude=0, longitude=0)
    
    # Preparar datos para el mapa
    puntos = []
    for c in clientes:
        puntos.append({
            'id': c.id,
            'nombre': c.nombre,
            'rut': c.rut,
            'medidor': c.medidor,
            'direccion': c.direccion,
            'lat': float(c.latitude),
            'lng': float(c.longitude),
        })
    
    context = {
        'empresa': empresa,
        'slug': alias,
        'puntos': json.dumps(puntos),
        'total_clientes': len(puntos),
    }
    return render(request, 'clientes/mapa_clientes.html', context)

# admin_ssr/views.py (o donde tengas las vistas de informes)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from empresas.models import Empresa
from clientes.models import Cliente
from datetime import datetime
import pytz

def escribir_cabecera_estandar(ws, empresa, usuario_nombre, fecha_hora, titulo_reporte, mes_anio=None):
    """
    Escribe la cabecera estándar en las primeras filas de la hoja Excel.
    Retorna el número de la siguiente fila disponible para comenzar a escribir datos.
    """
    row = 1
    font_normal = Font(name='Calibri', size=11)
    font_bold = Font(name='Calibri', size=11, bold=True)
    font_titulo = Font(name='Calibri', size=12, bold=True)
    alignment_left = Alignment(horizontal='left', vertical='center')

    # Fila 1: Nombre comité (mayúsculas, negrita)
    cell = ws.cell(row=row, column=1, value=empresa.nombre.upper())
    cell.font = font_bold
    cell.alignment = alignment_left
    row += 1

    # Fila 2: RUT (si el modelo Empresa tiene campo rut, si no se omite o se pone fijo)
    rut_empresa = getattr(empresa, 'rut', '') or ''
    if rut_empresa:
        cell = ws.cell(row=row, column=1, value=rut_empresa)
        cell.font = font_normal
        cell.alignment = alignment_left
        row += 1

    # Fila 3: Usuario
    cell = ws.cell(row=row, column=1, value=f"Usuario: {usuario_nombre}")
    cell.font = font_normal
    cell.alignment = alignment_left
    row += 1

    # Fila 4: Fecha y hora
    cell = ws.cell(row=row, column=1, value=fecha_hora)
    cell.font = font_normal
    cell.alignment = alignment_left
    row += 1

    # Fila 5: Título del reporte (mayúsculas, negrita)
    cell = ws.cell(row=row, column=1, value=titulo_reporte.upper())
    cell.font = font_titulo
    cell.alignment = alignment_left
    row += 1

    # Fila 6: Mes y año (opcional)
    if mes_anio:
        cell = ws.cell(row=row, column=1, value=mes_anio.upper())
        cell.font = font_normal
        cell.alignment = alignment_left
        row += 1

    row += 2  # Dejar una fila en blanco antes de los datos

    return row

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from empresas.models import Empresa
from clientes.models import Cliente, Contrato
from django.utils import timezone

def generar_pdf_contrato(request, alias, cliente_id):
    """
    Genera un PDF con los datos del contrato de un cliente específico.
    """
    db_alias = f'db_{alias}'
    empresa = get_object_or_404(Empresa, slug=alias)
    cliente = get_object_or_404(Cliente.objects.using(db_alias), id=cliente_id)
    try:
        contrato = cliente.contrato
    except Contrato.DoesNotExist:
        contrato = None

    # Respuesta PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="contrato_{cliente.id}.pdf"'

    # Documento en orientación horizontal
    doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                            rightMargin=72, leftMargin=72,
                            topMargin=72, bottomMargin=72)

    styles = getSampleStyleSheet()
    style_title = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=30,
        textColor=colors.HexColor('#059669')
    )
    style_heading = ParagraphStyle(
        'Heading2',
        parent=styles['Heading2'],
        fontSize=12,
        spaceAfter=12,
        textColor=colors.HexColor('#374151')
    )
    # Estilo para contenido de tabla con fuente más pequeña
    style_table = ParagraphStyle(
        'TableContent',
        parent=styles['Normal'],
        fontSize=9,
        leading=12
    )

    story = []

    # Título
    story.append(Paragraph(f"Informe de Contrato - {empresa.nombre}", style_title))
    story.append(Spacer(1, 0.25*inch))

    # ===================== DATOS DEL CLIENTE =====================
    story.append(Paragraph("Datos del Cliente", style_heading))

    # Lista de pares (etiqueta, valor)
    campos_cliente = [
        ("Nombre", cliente.nombre),
        ("RUT", cliente.rut),
        ("Apellido Paterno", cliente.apellido_paterno or '-'),
        ("Apellido Materno", cliente.apellido_materno or '-'),
        ("Dirección", cliente.direccion or '-'),
        ("Sector", cliente.sector or '-'),
        ("Comuna", getattr(cliente, 'comuna', '-')),
        ("Ciudad", getattr(cliente, 'ciudad', '-')),
        ("Teléfono", cliente.telefono or '-'),
        ("Email", cliente.email or '-'),
        ("Fecha Nacimiento", cliente.fecha_nacimiento.strftime('%d/%m/%Y') if cliente.fecha_nacimiento else '-'),
        ("Sexo", cliente.sexo or '-'),
        ("Estado Civil", cliente.estado_civil or '-'),
        ("Profesión", cliente.profesion or '-'),
        ("Fecha Incorporación", cliente.fecha_incorporacion.strftime('%d/%m/%Y') if cliente.fecha_incorporacion else '-'),
        ("Número Libro", cliente.numero_libro or '-'),
    ]

    # Construir tabla con 4 columnas (etiqueta1, valor1, etiqueta2, valor2)
    data_cliente = []
    for i in range(0, len(campos_cliente), 2):
        fila = []
        # Primera columna (par i)
        if i < len(campos_cliente):
            label = f"<b>{campos_cliente[i][0]}</b>"
            value = str(campos_cliente[i][1]) if campos_cliente[i][1] is not None else '-'
            fila.append(Paragraph(label, style_table))
            fila.append(Paragraph(value, style_table))
        else:
            fila.append("")
            fila.append("")
        # Segunda columna (par i+1)
        if i+1 < len(campos_cliente):
            label = f"<b>{campos_cliente[i+1][0]}</b>"
            value = str(campos_cliente[i+1][1]) if campos_cliente[i+1][1] is not None else '-'
            fila.append(Paragraph(label, style_table))
            fila.append(Paragraph(value, style_table))
        else:
            fila.append("")
            fila.append("")
        data_cliente.append(fila)

    # Anchos de columna (4 columnas)
    col_widths = [2.2*inch, 2.8*inch, 2.2*inch, 2.8*inch]  # Total ~10 pulgadas (A4 landscape mide 11.69")
    table_cliente = Table(data_cliente, colWidths=col_widths)
    table_cliente.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(table_cliente)
    story.append(Spacer(1, 0.25*inch))

    # ===================== DATOS DEL CONTRATO =====================
    if contrato:
        story.append(Paragraph("Datos del Contrato", style_heading))
        campos_contrato = [
            ("Tipo Cliente", contrato.tipo_cliente or '-'),
            ("Tipo Servicio SSR", contrato.tipo_servicio_ssr or '-'),
            ("Fecha Contrato", contrato.fecha_contrato.strftime('%d/%m/%Y') if contrato.fecha_contrato else '-'),
            ("Tarifa", contrato.tarifa or '-'),
            ("Diámetro", str(contrato.diametro) if contrato.diametro else '-'),
            ("Tipo Servicio", contrato.tipo_servicio or '-'),
            ("Marca Medidor", contrato.marca_medidor or '-'),
            ("Número Medidor", contrato.numero_medidor or '-'),
            ("Año Medidor", str(contrato.ano_medidor) if contrato.ano_medidor else '-'),
            ("Tipo Medidor", contrato.tipo_medidor or '-'),
            ("Sello Medidor", contrato.sello_medidor or '-'),
            ("Código Unión Domiciliaria", contrato.codigo_union_domiciliaria or '-'),
            ("UTM Norte", contrato.utm_norte or '-'),
            ("UTM Este", contrato.utm_este or '-'),
            ("Email Recepción Documento", contrato.email_recepcion_documento or '-'),
            ("Socio", "Sí" if contrato.socio else "No"),
        ]

        data_contrato = []
        for i in range(0, len(campos_contrato), 1):
            fila = []
            if i < len(campos_contrato):
                label = f"<b>{campos_contrato[i][0]}</b>"
                value = str(campos_contrato[i][1]) if campos_contrato[i][1] is not None else '-'
                fila.append(Paragraph(label, style_table))
                fila.append(Paragraph(value, style_table))
            else:
                fila.append("")
                fila.append("")
            if i+1 < len(campos_contrato):
                label = f"<b>{campos_contrato[i+1][0]}</b>"
                value = str(campos_contrato[i+1][1]) if campos_contrato[i+1][1] is not None else '-'
                fila.append(Paragraph(label, style_table))
                fila.append(Paragraph(value, style_table))
            else:
                fila.append("")
                fila.append("")
            data_contrato.append(fila)

        table_contrato = Table(data_contrato, colWidths=col_widths)
        table_contrato.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('PADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(table_contrato)
    else:
        story.append(Paragraph("No existe contrato asociado", styles['Normal']))

    # Pie de página
    

    doc.build(story)
    return response

def informe_contratos(request, alias):
    """Página para generar el informe de contratos con opciones de ordenamiento."""
    empresa = get_object_or_404(Empresa, slug=alias)
    return render(request, 'informe_contratos.html', {
        'empresa': empresa,
        'slug': alias,
    })

def informe_socios(request, alias):
    empresa = get_object_or_404(Empresa, slug=alias)
    return render(request, 'informe_socios.html', {
        'empresa': empresa,
        'slug': alias,
    })

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from empresas.models import Empresa
from clientes.models import Cliente
from datetime import datetime
import pytz

def exportar_contratos_excel(request, alias):
    db_alias = f'db_{alias}'
    empresa = get_object_or_404(Empresa, slug=alias)

    user = request.user
    usuario_nombre = user.get_full_name() or user.username if user.is_authenticated else "Anónimo"
    santiago_tz = pytz.timezone('America/Santiago')
    ahora = datetime.now(santiago_tz)
    fecha_hora = ahora.strftime('%d/%m/%Y %H:%M:%S')
    fecha_archivo = ahora.strftime('%Y%m%d_%H%M%S')

    meses_es = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
    mes_anio = f"{meses_es[ahora.month-1]} de {ahora.year}"

    order_by = request.GET.get('order_by', 'apellido_paterno')
    direction = request.GET.get('direction', 'asc')
    if direction == 'desc':
        order_by = f'-{order_by}'
    allowed_fields = ['apellido_paterno', 'fecha_incorporacion', 'sector']
    if order_by.lstrip('-') not in allowed_fields:
        order_by = 'apellido_paterno'

    clientes = Cliente.objects.using(db_alias).select_related('contrato').order_by(order_by)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contratos"

    # Ocultar líneas de cuadrícula para un aspecto más limpio
    ws.sheet_view.showGridLines = False

    titulo_reporte = "LISTADO DE CONTRATOS"
    fila_inicio_tabla = escribir_cabecera_estandar(
        ws, empresa, usuario_nombre, fecha_hora, titulo_reporte, mes_anio
    )

    # --- Estilos de tabla (tradicional) ---
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    data_font = Font(name='Calibri', size=9)
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_alignment_right = Alignment(horizontal='right', vertical='center')

    # Bordes completos (clásico)
    full_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # --- Encabezados ---
    headers = [
        'Número', 'Sector', 'Ruta', 'Dirección', 'Comuna', 'Ciudad', 'Rol',
        'Tipo Cliente', 'Contrato', 'Servicio', 'RUT', 'Nombre', 'Apellido Paterno',
        'Apellido Materno', 'Sexo', 'Estado Civil', 'Fecha Nacimiento',
        'Profesión / Oficio', 'Fecha Defunción', 'Fecha Contrato', 'Fecha Incorporación',
        'Número Libro', 'Fono 1', 'Fono 2', 'Email Contacto', 'Email Recepción Documento',
        'Tarifa', 'Diámetro', 'Tipo Servicio', 'Tipo SSR', 'Subsidio', 'Número Medidor',
        'Marca Medidor', 'Año Medidor', 'Sello Medidor', 'Tipo Medidor',
        'Código Unión Domiciliaria', 'UTM Norte', 'UTM Este', 'Documento', 'Razón Social', 'Socio'
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=fila_inicio_tabla, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = full_border

    # --- Datos ---
    for row_idx, cliente in enumerate(clientes, start=fila_inicio_tabla + 1):
        contrato = getattr(cliente, 'contrato', None)

        row_data = [
            cliente.id,
            cliente.sector or '',
            getattr(cliente, 'ruta', '') or '',
            cliente.direccion or '',
            getattr(contrato, 'comuna', '') or '',
            getattr(contrato, 'ciudad', '') or '',
            getattr(contrato, 'rol', '') or '',
            contrato.tipo_cliente if contrato else '',
            getattr(contrato, 'numero_contrato', '') or '',
            getattr(contrato, 'servicio', '') or '',
            cliente.rut or '',
            cliente.nombre or '',
            cliente.apellido_paterno or '',
            cliente.apellido_materno or '',
            cliente.sexo or '',
            cliente.estado_civil or '',
            cliente.fecha_nacimiento.strftime('%d/%m/%Y') if cliente.fecha_nacimiento else '',
            cliente.profesion or '',
            cliente.fecha_defuncion.strftime('%d/%m/%Y') if cliente.fecha_defuncion else '',
            contrato.fecha_contrato.strftime('%d/%m/%Y') if contrato and contrato.fecha_contrato else '',
            cliente.fecha_incorporacion.strftime('%d/%m/%Y') if cliente.fecha_incorporacion else '',
            cliente.numero_libro or '',
            cliente.telefono or '',
            cliente.contacto2 or '',
            cliente.email_contacto or '',
            contrato.email_recepcion_documento if contrato else '',
            contrato.tarifa if contrato else '',
            contrato.diametro if contrato else '',
            contrato.tipo_servicio if contrato else '',
            contrato.tipo_servicio_ssr if contrato else '',
            getattr(contrato, 'subsidio', '') or '',
            cliente.medidor or '',
            contrato.marca_medidor if contrato else '',
            contrato.ano_medidor if contrato else '',
            contrato.sello_medidor if contrato else '',
            contrato.tipo_medidor if contrato else '',
            contrato.codigo_union_domiciliaria if contrato else '',
            contrato.utm_norte if contrato else '',
            contrato.utm_este if contrato else '',
            getattr(contrato, 'documento', '') or '',
            getattr(cliente, 'razon_social', '') or '',
            'Sí' if contrato and contrato.socio else 'No',
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment_right if isinstance(value, (int, float)) else data_alignment
            cell.border = full_border
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                cell.number_format = '#,##0.00'
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Ajustar ancho de columnas
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        for row in range(fila_inicio_tabla, ws.max_row + 1):
            cell_value = ws.cell(row, col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 4, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Congelar paneles
    ws.freeze_panes = f'A{fila_inicio_tabla + 1}'

    # Pie de página
    last_row = ws.max_row + 1
    ws[f'A{last_row}'] = f"Reporte generado el {ahora.strftime('%d/%m/%Y %H:%M:%S')}"
    ws[f'A{last_row}'].font = Font(name='Calibri', size=8, italic=True, color='7F8C8D')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"contratos_{alias}_{fecha_archivo}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from empresas.models import Empresa
from clientes.models import Cliente
from datetime import datetime
import pytz

def exportar_socios_excel(request, alias):
    db_alias = f'db_{alias}'
    empresa = get_object_or_404(Empresa, slug=alias)

    user = request.user
    usuario_nombre = user.get_full_name() or user.username if user.is_authenticated else "Anónimo"
    santiago_tz = pytz.timezone('America/Santiago')
    ahora = datetime.now(santiago_tz)
    fecha_hora = ahora.strftime('%d/%m/%Y %H:%M:%S')
    fecha_archivo = ahora.strftime('%Y%m%d_%H%M%S')

    meses_es = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
    mes_anio = f"{meses_es[ahora.month-1]} de {ahora.year}"

    order_by = request.GET.get('order_by', 'apellido_paterno')
    direction = request.GET.get('direction', 'asc')
    if direction == 'desc':
        order_by = f'-{order_by}'

    allowed_fields = ['apellido_paterno', 'fecha_incorporacion', 'sector']
    if order_by.lstrip('-') not in allowed_fields:
        order_by = 'apellido_paterno'

    clientes = Cliente.objects.using(db_alias).select_related('contrato').order_by(order_by)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Socios"

    # Ocultar líneas de cuadrícula
    ws.sheet_view.showGridLines = False

    titulo_reporte = "LISTADO DE SOCIOS"
    fila_inicio_tabla = escribir_cabecera_estandar(
        ws, empresa, usuario_nombre, fecha_hora, titulo_reporte, mes_anio
    )

    # --- Estilos de tabla (tradicional) ---
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    data_font = Font(name='Calibri', size=9)
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_alignment_right = Alignment(horizontal='right', vertical='center')

    full_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # --- Encabezados ---
    headers = [
        'Sector', 'Ruta', 'N° Libro', 'Contrato', 'RUT',
        'Nombre', 'Apellido Paterno', 'Apellido Materno',
        'Fec. Incorporación', 'Fec. Contrato', 'Sexo',
        'Dirección Arranque', 'Dirección', 'Servicio', '¿Socio?'
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=fila_inicio_tabla, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = full_border

    # --- Datos ---
    for row_idx, cliente in enumerate(clientes, start=fila_inicio_tabla + 1):
        contrato = cliente.contrato if hasattr(cliente, 'contrato') else None

        nombre_completo = cliente.nombre or ''
        if cliente.apellido_paterno:
            nombre_completo += ' ' + cliente.apellido_paterno
        if cliente.apellido_materno:
            nombre_completo += ' ' + cliente.apellido_materno
        nombre_completo = nombre_completo.strip() or '-'

        row_data = [
            cliente.sector or '',
            getattr(cliente, 'ruta', '') or '',
            cliente.numero_libro or '',
            getattr(contrato, 'numero_contrato', '') if contrato else '',
            cliente.rut or '',
            nombre_completo,
            cliente.apellido_paterno or '',
            cliente.apellido_materno or '',
            cliente.fecha_incorporacion.strftime('%d/%m/%Y') if cliente.fecha_incorporacion else '',
            contrato.fecha_contrato.strftime('%d/%m/%Y') if contrato and contrato.fecha_contrato else '',
            cliente.sexo or '',
            contrato.direccion_arranque if contrato else '',
            cliente.direccion or '',
            contrato.servicio if contrato else '',
            'Sí' if (contrato and contrato.socio) else 'No'
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment_right if isinstance(value, (int, float)) else data_alignment
            cell.border = full_border
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                cell.number_format = '#,##0.00'
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Ajustar ancho de columnas
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        for row in range(fila_inicio_tabla, ws.max_row + 1):
            cell_value = ws.cell(row, col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 4, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Congelar paneles
    ws.freeze_panes = f'A{fila_inicio_tabla + 1}'

    # Pie de página
    last_row = ws.max_row + 1
    ws[f'A{last_row}'] = f"Reporte generado el {ahora.strftime('%d/%m/%Y %H:%M:%S')}"
    ws[f'A{last_row}'].font = Font(name='Calibri', size=8, italic=True, color='7F8C8D')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"socios_{alias}_{fecha_archivo}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

from datetime import datetime
from decimal import Decimal
import os
from django.conf import settings
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404

@login_required
@require_POST
@permiso_requerido('cambio_medidor')
def registrar_cambio_medidor_ajax(request, alias, cliente_id):
    db_alias = f'db_{alias}'
    db_path = os.path.join(settings.BASES_DIR, f'{db_alias}.sqlite3')
    if not os.path.exists(db_path):
        return JsonResponse({'success': False, 'error': 'Base de datos no encontrada'}, status=404)

    try:
        from empresas.models import Empresa
        from clientes.models import Cliente, CambioMedidor
        from lecturas.models import LecturaMovil

        empresa = get_object_or_404(Empresa, slug=alias)
        cliente = get_object_or_404(Cliente.objects.using(db_alias), id=cliente_id)

        # --- Obtener datos del POST ---
        fecha_lectura_anterior_str = request.POST.get('fecha_lectura_anterior')
        lectura_anterior_str = request.POST.get('lectura_anterior')

        # --- Intentar obtener automáticamente de la última lectura si no se enviaron ---
        ultima_lectura = None
        if not fecha_lectura_anterior_str or not lectura_anterior_str:
            ultima_lectura = LecturaMovil.objects.filter(
                empresa_slug=alias,
                cliente=cliente_id,
                consumo__isnull=False,
                consumo__gt=0          # Solo lecturas con consumo positivo
            ).order_by('-fecha_lectura').first()

            if ultima_lectura:
                if not fecha_lectura_anterior_str:
                    fecha_lectura_anterior_str = ultima_lectura.fecha_lectura.strftime('%Y-%m-%d')
                if not lectura_anterior_str:
                    lectura_anterior_str = str(ultima_lectura.lectura_actual)

        # --- Validaciones ---
        lectura_retiro_str = request.POST.get('lectura_retiro')
        if not lectura_retiro_str:
            return JsonResponse({'success': False, 'error': 'La lectura de retiro es obligatoria'}, status=400)

        fecha_instalacion_str = request.POST.get('fecha_instalacion')
        if not fecha_instalacion_str:
            return JsonResponse({'success': False, 'error': 'Fecha de instalación requerida'})

        # --- Conversión a tipos Python ---
        # Fecha lectura anterior (puede ser None)
        fecha_lectura_anterior = None
        if fecha_lectura_anterior_str:
            fecha_lectura_anterior = datetime.strptime(fecha_lectura_anterior_str, '%Y-%m-%d').date()

        # Fecha instalación
        fecha_instalacion = datetime.strptime(fecha_instalacion_str, '%Y-%m-%d').date()

        # Lecturas a Decimal
        lectura_anterior_dec = Decimal(lectura_anterior_str) if lectura_anterior_str else Decimal('0')
        lectura_retiro_dec = Decimal(lectura_retiro_str)
        lectura_inicial_dec = Decimal(request.POST.get('lectura_inicial', '0'))

        # --- Crear registro ---
        cambio = CambioMedidor.objects.using(db_alias).create(
            cliente=cliente,
            medidor_retirado_marca=request.POST.get('medidor_retirado_marca', ''),
            medidor_retirado_numero=request.POST.get('medidor_retirado_numero', ''),
            medidor_retirado_anio=request.POST.get('medidor_retirado_anio'),
            fecha_lectura_anterior=fecha_lectura_anterior,
            lectura_anterior=lectura_anterior_dec,
            lectura_retiro=lectura_retiro_dec,
            medidor_nuevo_marca=request.POST.get('medidor_nuevo_marca', ''),
            medidor_nuevo_numero=request.POST.get('medidor_nuevo_numero', ''),
            medidor_nuevo_anio=request.POST.get('medidor_nuevo_anio'),
            fecha_instalacion=fecha_instalacion,
            lectura_inicial=lectura_inicial_dec,
            usuario=request.user.username if request.user.is_authenticated else 'Sistema'
        )

        # El método save() del modelo ya actualiza cliente.medidor y contrato

        return JsonResponse({
            'success': True,
            'cambio': {
                'id': cambio.id,
                'periodo': cambio.periodo or '-',
                'marca_retirado': cambio.medidor_retirado_marca or '-',
                'numero_retirado': cambio.medidor_retirado_numero,
                'ano_retirado': cambio.medidor_retirado_anio or '-',
                'fecha_lectura_anterior': cambio.fecha_lectura_anterior.strftime('%d/%m/%Y') if cambio.fecha_lectura_anterior else '-',
                'lectura_anterior': float(cambio.lectura_anterior) if cambio.lectura_anterior else 0,
                'lectura_retiro': float(cambio.lectura_retiro),
                'consumo': float(cambio.consumo_final) if cambio.consumo_final else 0,
                'fecha_instalacion': cambio.fecha_instalacion.strftime('%d/%m/%Y'),
                'numero_instalado': cambio.medidor_nuevo_numero,
                'lectura_inicial': float(cambio.lectura_inicial),
                'fecha_registro': cambio.fecha_registro.strftime('%d/%m/%Y %H:%M'),
                'usuario': cambio.usuario or '-',
            }
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
import json

# -------------------------------------------------------------------
# 1. SUBSIDIO
# -------------------------------------------------------------------
@login_required
@require_POST
@permiso_requerido('subsidio') 
def crear_subsidio_ajax(request, alias, cliente_id):
    db_alias = f'db_{alias}'
    cliente = get_object_or_404(Cliente.objects.using(db_alias), id=cliente_id)

    try:
        Subsidio.objects.using(db_alias).create(
            cliente=cliente,
            subsidio=request.POST.get('subsidio'),
            municipalidad=request.POST.get('municipalidad'),
            numero_decreto=request.POST.get('numero_decreto'),
            tramo=request.POST.get('tramo'),
            rut_beneficiario=request.POST.get('rut_beneficiario'),
            beneficiario=request.POST.get('beneficiario'),
            vivienda=request.POST.get('vivienda', ''),
            fecha_decreto=request.POST.get('fecha_decreto'),
            fecha_inicio=request.POST.get('fecha_inicio'),
            fecha_vencimiento=request.POST.get('fecha_vencimiento'),
            fecha_encuesta=request.POST.get('fecha_encuesta') or None,
        )
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# -------------------------------------------------------------------
# 2. CARGO PERMANENTE
# -------------------------------------------------------------------
@login_required
@require_POST
@permiso_requerido('cargo_permanente')
def crear_cargo_permanente_ajax(request, alias, cliente_id):
    db_alias = f'db_{alias}'
    cliente = get_object_or_404(Cliente.objects.using(db_alias), id=cliente_id)

    try:
        CargoPermanente.objects.using(db_alias).create(
            cliente=cliente,
            fecha=request.POST.get('fecha'),
            codigo=request.POST.get('codigo'),
            descripcion=request.POST.get('descripcion'),
            monto=request.POST.get('monto'),
            facturable=request.POST.get('facturable') == 'on',
            periodo=request.POST.get('periodo'),
            usuario=request.user.username if request.user.is_authenticated else 'Sistema',
        )
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# -------------------------------------------------------------------
# 3. CARGO
# -------------------------------------------------------------------
@login_required
@require_POST
@permiso_requerido('cargo')
def crear_cargo_ajax(request, alias, cliente_id):
    db_alias = f'db_{alias}'
    cliente = get_object_or_404(Cliente.objects.using(db_alias), id=cliente_id)

    try:
        Cargo.objects.using(db_alias).create(
            cliente=cliente,
            fecha=request.POST.get('fecha'),
            codigo=request.POST.get('codigo'),
            descripcion=request.POST.get('descripcion'),
            monto=request.POST.get('monto'),
            facturable=request.POST.get('facturable') == 'on',
            periodo=request.POST.get('periodo'),
            usuario=request.user.username if request.user.is_authenticated else 'Sistema',
        )
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# -------------------------------------------------------------------
# 4. DESCUENTO
# -------------------------------------------------------------------
@login_required
@require_POST
@permiso_requerido('descuento') 
def crear_descuento_ajax(request, alias, cliente_id):
    db_alias = f'db_{alias}'
    cliente = get_object_or_404(Cliente.objects.using(db_alias), id=cliente_id)

    try:
        Descuento.objects.using(db_alias).create(
            cliente=cliente,
            fecha=request.POST.get('fecha'),
            codigo=request.POST.get('codigo'),
            descripcion=request.POST.get('descripcion'),
            monto=request.POST.get('monto'),
            facturable=request.POST.get('facturable') == 'on',
            periodo=request.POST.get('periodo'),
            usuario=request.user.username if request.user.is_authenticated else 'Sistema',
        )
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# -------------------------------------------------------------------
# 5. CONVENIO
# -------------------------------------------------------------------
@login_required
@require_POST
@permiso_requerido('convenio')
def crear_convenio_ajax(request, alias, cliente_id):
    db_alias = f'db_{alias}'
    cliente = get_object_or_404(Cliente.objects.using(db_alias), id=cliente_id)

    try:
        Convenio.objects.using(db_alias).create(
            cliente=cliente,
            convenio=request.POST.get('convenio'),
            descripcion=request.POST.get('descripcion', ''),
            fecha_convenio=request.POST.get('fecha_convenio'),
            total_cuotas=request.POST.get('total_cuotas'),
            monto=request.POST.get('monto'),
            facturable=request.POST.get('facturable') == 'on',
            usuario=request.user.username if request.user.is_authenticated else 'Sistema',
        )
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# -------------------------------------------------------------------
# 6. OTRO INGRESO
# -------------------------------------------------------------------
@login_required
@require_POST
@permiso_requerido('otro_ingreso')
def crear_otro_ingreso_ajax(request, alias, cliente_id):
    db_alias = f'db_{alias}'
    cliente = get_object_or_404(Cliente.objects.using(db_alias), id=cliente_id)

    try:
        OtroIngreso.objects.using(db_alias).create(
            cliente=cliente,
            concepto=request.POST.get('concepto'),
            documento=request.POST.get('documento', ''),
            fecha_pago=request.POST.get('fecha_pago'),
            monto=request.POST.get('monto'),
            comprobante=request.POST.get('comprobante', ''),
        )
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# -------------------------------------------------------------------
# 7. CORTE Y REPOSICIÓN
# -------------------------------------------------------------------
@login_required
@require_POST
@permiso_requerido('corte')
def crear_corte_reposicion_ajax(request, alias, cliente_id):
    db_alias = f'db_{alias}'
    cliente = get_object_or_404(Cliente.objects.using(db_alias), id=cliente_id)

    try:
        CorteReposicion.objects.using(db_alias).create(
            cliente=cliente,
            fecha_corte=request.POST.get('fecha_corte'),
            operador_corte=request.POST.get('operador_corte'),
            lectura_corte=request.POST.get('lectura_corte'),
            tipo_corte=request.POST.get('tipo_corte'),
        )
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)



import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from datetime import datetime
import pytz
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

# Asegúrate de importar tus modelos
from clientes.models import (
    Subsidio, CargoPermanente, Cargo, Descuento,
    Convenio, OtroIngreso, CorteReposicion
)


def generar_reporte_generico(request, alias, modelo, titulo, headers, campos, nombre_archivo):
    """
    Si no se especifica formato (?formato=), muestra página de selección.
    Si se especifica, genera el archivo solicitado.
    """
    db_alias = f'db_{alias}'
    empresa = get_object_or_404(Empresa, slug=alias)
    
    formato = request.GET.get('formato')          # ← ya no ponemos valor por defecto
    

    if not formato:
        context = {
            'empresa': empresa,
            'slug': alias,
            'titulo': titulo,
            'nombre_archivo': nombre_archivo,
            'url_base': request.path,              # la misma URL sin ?formato
        }
        return render(request, 'seleccionar_formato.html', context)
    # ------------------------------------------------
    

    objetos = modelo.objects.using(db_alias).all().order_by('-id')
    
    user = request.user
    usuario_nombre = user.get_full_name() or user.username if user.is_authenticated else "Anónimo"
    santiago_tz = pytz.timezone('America/Santiago')
    ahora = datetime.now(santiago_tz)
    fecha_hora = ahora.strftime('%d/%m/%Y %H:%M:%S')
    fecha_archivo = ahora.strftime('%Y%m%d_%H%M%S')
    mes_anio = ahora.strftime('%B de %Y').upper()
    
    if formato == 'pdf':
        return generar_pdf_generico(empresa, usuario_nombre, fecha_hora, titulo, mes_anio, headers, objetos, campos, nombre_archivo)
    else:
        return generar_excel_generico(empresa, usuario_nombre, fecha_hora, titulo, mes_anio, headers, objetos, campos, nombre_archivo, fecha_archivo)


def generar_excel_generico(empresa, usuario_nombre, fecha_hora, titulo, subtitulo, headers, objetos, campos, nombre_archivo, fecha_archivo):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = nombre_archivo[:30]
    ws.sheet_view.showGridLines = False

    # Cabecera estándar
    fila_inicio_tabla = escribir_cabecera_estandar(
        ws, empresa, usuario_nombre, fecha_hora, titulo, subtitulo
    )

    # Estilos
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='Calibri', size=9)
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_alignment_right = Alignment(horizontal='right', vertical='center')
    full_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # Encabezados
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=fila_inicio_tabla, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = full_border

    # Datos
    for row_idx, obj in enumerate(objetos, start=fila_inicio_tabla + 1):
        for col_idx, campo in enumerate(campos, 1):
            valor = getattr(obj, campo, '')
            if isinstance(valor, datetime):
                valor = valor.strftime('%d/%m/%Y')
            elif isinstance(valor, bool):
                valor = 'Sí' if valor else 'No'
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = data_font
            cell.alignment = data_alignment_right if isinstance(valor, (int, float)) else data_alignment
            cell.border = full_border
            if isinstance(valor, (int, float)):
                cell.number_format = '#,##0.00'
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Ajustar ancho de columnas
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        for row in range(fila_inicio_tabla, ws.max_row + 1):
            val = ws.cell(row, col_idx).value
            if val:
                max_length = max(max_length, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 3, 50)

    ws.freeze_panes = f'A{fila_inicio_tabla + 1}'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}_{empresa.slug}_{fecha_archivo}.xlsx"'
    wb.save(response)
    return response


def generar_pdf_generico(empresa, usuario_nombre, fecha_hora, titulo, subtitulo, headers, objetos, campos, nombre_archivo):
    response = HttpResponse(content_type='application/pdf')
    # Cambiar a 'attachment' para que descargue con nombre correcto
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}_{empresa.slug}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                            rightMargin=30, leftMargin=30,
                            topMargin=30, bottomMargin=30)
    
    # 🔥 Establecer metadatos del PDF (aparecerá como título de la pestaña)
    doc.title = titulo

    styles = getSampleStyleSheet()
    style_title = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=14, alignment=1, spaceAfter=10)
    style_subtitle = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=1, spaceAfter=20)

    story = []
    story.append(Paragraph(f"{titulo} - {empresa.nombre.upper()}", style_title))
    story.append(Paragraph(subtitulo, style_subtitle))
    story.append(Paragraph(f"Generado por: {usuario_nombre} | {fecha_hora}", styles['Normal']))
    story.append(Spacer(1, 0.2*inch))

    # Tabla (igual que antes)
    data = [headers]
    for obj in objetos:
        fila = []
        for campo in campos:
            valor = getattr(obj, campo, '')
            if isinstance(valor, datetime):
                valor = valor.strftime('%d/%m/%Y')
            elif isinstance(valor, bool):
                valor = 'Sí' if valor else 'No'
            fila.append(str(valor))
        data.append(fila)

    if not objetos:
        data.append(['No hay registros para el período seleccionado'] + [''] * (len(headers)-1))

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2E75B6')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F2F2F2')]),
    ]))
    story.append(table)

    doc.build(story)
    return response

def informe_subsidios(request, alias):
    headers = ['Subsidio', 'Municipalidad', 'N° Decreto', 'Tramo', 'RUT Beneficiario', 'Beneficiario', 'Vivienda', 'Fecha Decreto', 'Fecha Inicio', 'Fecha Vencimiento', 'Fecha Encuesta']
    campos = ['subsidio', 'municipalidad', 'numero_decreto', 'tramo', 'rut_beneficiario', 'beneficiario', 'vivienda', 'fecha_decreto', 'fecha_inicio', 'fecha_vencimiento', 'fecha_encuesta']
    return generar_reporte_generico(request, alias, Subsidio, "LISTADO DE SUBSIDIOS", headers, campos, "subsidios")

def informe_cargos_permanentes(request, alias):
    headers = ['Fecha', 'Código', 'Descripción', 'Monto', 'Facturable', 'Periodo', 'Usuario']
    campos = ['fecha', 'codigo', 'descripcion', 'monto', 'facturable', 'periodo', 'usuario']
    return generar_reporte_generico(request, alias, CargoPermanente, "LISTADO DE CARGOS PERMANENTES", headers, campos, "cargos_permanentes")

def informe_cargos(request, alias):
    headers = ['Fecha', 'Código', 'Descripción', 'Monto', 'Facturable', 'Periodo', 'Usuario']
    campos = ['fecha', 'codigo', 'descripcion', 'monto', 'facturable', 'periodo', 'usuario']
    return generar_reporte_generico(request, alias, Cargo, "LISTADO DE CARGOS", headers, campos, "cargos")

def informe_descuentos(request, alias):
    headers = ['Fecha', 'Código', 'Descripción', 'Monto', 'Facturable', 'Periodo', 'Usuario']
    campos = ['fecha', 'codigo', 'descripcion', 'monto', 'facturable', 'periodo', 'usuario']
    return generar_reporte_generico(request, alias, Descuento, "LISTADO DE DESCUENTOS", headers, campos, "descuentos")

def informe_convenios(request, alias):
    headers = ['Convenio', 'Descripción', 'Fecha Convenio', 'Total Cuotas', 'Monto', 'Facturable', 'Usuario']
    campos = ['convenio', 'descripcion', 'fecha_convenio', 'total_cuotas', 'monto', 'facturable', 'usuario']
    return generar_reporte_generico(request, alias, Convenio, "LISTADO DE CONVENIOS", headers, campos, "convenios")

def informe_otros_ingresos(request, alias):
    headers = ['Concepto', 'Documento', 'Fecha Pago', 'Monto', 'Comprobante']
    campos = ['concepto', 'documento', 'fecha_pago', 'monto', 'comprobante']
    return generar_reporte_generico(request, alias, OtroIngreso, "LISTADO DE OTROS INGRESOS", headers, campos, "otros_ingresos")

def informe_cortes(request, alias):
    headers = ['Fecha Corte', 'Operador', 'Lectura Corte', 'Tipo Corte', 'Fecha Reposición', 'Operador Reposición', 'Anulado']
    campos = ['fecha_corte', 'operador_corte', 'lectura_corte', 'tipo_corte', 'fecha_reposicion', 'operador_reposicion', 'anulado']
    return generar_reporte_generico(request, alias, CorteReposicion, "LISTADO DE CORTES Y REPOSICIONES", headers, campos, "cortes")

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

def _build_pdf(titulo, empresa, usuario_nombre, fecha_hora, mes_anio,
               headers, data_rows, filename):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                            rightMargin=30, leftMargin=30,
                            topMargin=30, bottomMargin=30)
    doc.title = titulo   # metadato para el nombre en la pestaña

    styles = getSampleStyleSheet()
    style_title = ParagraphStyle('Title', parent=styles['Heading1'],
                                 fontSize=14, alignment=1, spaceAfter=10)
    style_subtitle = ParagraphStyle('Subtitle', parent=styles['Normal'],
                                    fontSize=10, alignment=1, spaceAfter=20)
    style_normal = styles['Normal']

    story = []
    story.append(Paragraph(f"{titulo} - {empresa.nombre.upper()}", style_title))
    story.append(Paragraph(mes_anio, style_subtitle))
    story.append(Paragraph(f"Generado por: {usuario_nombre} | {fecha_hora}", style_normal))
    story.append(Spacer(1, 0.2*inch))

    # Tabla
    table_data = [headers] + data_rows
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2E75B6')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F2F2F2')]),
    ]))
    story.append(table)

    doc.build(story)
    return response

def exportar_contratos_pdf(request, alias):
    db_alias = f'db_{alias}'
    empresa = get_object_or_404(Empresa, slug=alias)

    user = request.user
    usuario_nombre = user.get_full_name() or user.username if user.is_authenticated else "Anónimo"
    santiago_tz = pytz.timezone('America/Santiago')
    ahora = datetime.now(santiago_tz)
    fecha_hora = ahora.strftime('%d/%m/%Y %H:%M:%S')
    fecha_archivo = ahora.strftime('%Y%m%d_%H%M%S')

    meses_es = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
    mes_anio = f"{meses_es[ahora.month-1]} de {ahora.year}"

    order_by = request.GET.get('order_by', 'apellido_paterno')
    direction = request.GET.get('direction', 'asc')
    if direction == 'desc':
        order_by = f'-{order_by}'
    allowed_fields = ['apellido_paterno', 'fecha_incorporacion', 'sector']
    if order_by.lstrip('-') not in allowed_fields:
        order_by = 'apellido_paterno'

    clientes = Cliente.objects.using(db_alias).select_related('contrato').order_by(order_by)

    # Cabeceras (las mismas que en Excel)
    headers = [
        'Número', 'Sector', 'Ruta', 'Dirección', 'Comuna', 'Ciudad', 'Rol',
        'Tipo Cliente', 'Contrato', 'Servicio', 'RUT', 'Nombre', 'Apellido Paterno',
        'Apellido Materno', 'Sexo', 'Estado Civil', 'Fecha Nacimiento',
        'Profesión / Oficio', 'Fecha Defunción', 'Fecha Contrato', 'Fecha Incorporación',
        'Número Libro', 'Fono 1', 'Fono 2', 'Email Contacto', 'Email Recepción Documento',
        'Tarifa', 'Diámetro', 'Tipo Servicio', 'Tipo SSR', 'Subsidio', 'Número Medidor',
        'Marca Medidor', 'Año Medidor', 'Sello Medidor', 'Tipo Medidor',
        'Código Unión Domiciliaria', 'UTM Norte', 'UTM Este', 'Documento', 'Razón Social', 'Socio'
    ]

    data_rows = []
    for cliente in clientes:
        contrato = getattr(cliente, 'contrato', None)
        row = [
            cliente.id,
            cliente.sector or '',
            getattr(cliente, 'ruta', '') or '',
            cliente.direccion or '',
            getattr(contrato, 'comuna', '') or '',
            getattr(contrato, 'ciudad', '') or '',
            getattr(contrato, 'rol', '') or '',
            contrato.tipo_cliente if contrato else '',
            getattr(contrato, 'numero_contrato', '') or '',
            getattr(contrato, 'servicio', '') or '',
            cliente.rut or '',
            cliente.nombre or '',
            cliente.apellido_paterno or '',
            cliente.apellido_materno or '',
            cliente.sexo or '',
            cliente.estado_civil or '',
            cliente.fecha_nacimiento.strftime('%d/%m/%Y') if cliente.fecha_nacimiento else '',
            cliente.profesion or '',
            cliente.fecha_defuncion.strftime('%d/%m/%Y') if cliente.fecha_defuncion else '',
            contrato.fecha_contrato.strftime('%d/%m/%Y') if contrato and contrato.fecha_contrato else '',
            cliente.fecha_incorporacion.strftime('%d/%m/%Y') if cliente.fecha_incorporacion else '',
            cliente.numero_libro or '',
            cliente.telefono or '',
            cliente.contacto2 or '',
            cliente.email_contacto or '',
            contrato.email_recepcion_documento if contrato else '',
            contrato.tarifa if contrato else '',
            contrato.diametro if contrato else '',
            contrato.tipo_servicio if contrato else '',
            contrato.tipo_servicio_ssr if contrato else '',
            getattr(contrato, 'subsidio', '') or '',
            cliente.medidor or '',
            contrato.marca_medidor if contrato else '',
            contrato.ano_medidor if contrato else '',
            contrato.sello_medidor if contrato else '',
            contrato.tipo_medidor if contrato else '',
            contrato.codigo_union_domiciliaria if contrato else '',
            contrato.utm_norte if contrato else '',
            contrato.utm_este if contrato else '',
            getattr(contrato, 'documento', '') or '',
            getattr(cliente, 'razon_social', '') or '',
            'Sí' if contrato and contrato.socio else 'No',
        ]
        data_rows.append(row)

    titulo = "LISTADO DE CONTRATOS"
    filename = f"contratos_{alias}_{fecha_archivo}.pdf"
    return _build_pdf(titulo, empresa, usuario_nombre, fecha_hora, mes_anio,
                      headers, data_rows, filename)

def exportar_socios_pdf(request, alias):
    db_alias = f'db_{alias}'
    empresa = get_object_or_404(Empresa, slug=alias)

    user = request.user
    usuario_nombre = user.get_full_name() or user.username if user.is_authenticated else "Anónimo"
    santiago_tz = pytz.timezone('America/Santiago')
    ahora = datetime.now(santiago_tz)
    fecha_hora = ahora.strftime('%d/%m/%Y %H:%M:%S')
    fecha_archivo = ahora.strftime('%Y%m%d_%H%M%S')

    meses_es = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
    mes_anio = f"{meses_es[ahora.month-1]} de {ahora.year}"

    order_by = request.GET.get('order_by', 'apellido_paterno')
    direction = request.GET.get('direction', 'asc')
    if direction == 'desc':
        order_by = f'-{order_by}'
    allowed_fields = ['apellido_paterno', 'fecha_incorporacion', 'sector']
    if order_by.lstrip('-') not in allowed_fields:
        order_by = 'apellido_paterno'

    clientes = Cliente.objects.using(db_alias).select_related('contrato').order_by(order_by)

    headers = [
        'Sector', 'Ruta', 'N° Libro', 'Contrato', 'RUT',
        'Nombre', 'Apellido Paterno', 'Apellido Materno',
        'Fec. Incorporación', 'Fec. Contrato', 'Sexo',
        'Dirección Arranque', 'Dirección', 'Servicio', '¿Socio?'
    ]

    data_rows = []
    for cliente in clientes:
        contrato = cliente.contrato if hasattr(cliente, 'contrato') else None
        nombre_completo = cliente.nombre or ''
        if cliente.apellido_paterno:
            nombre_completo += ' ' + cliente.apellido_paterno
        if cliente.apellido_materno:
            nombre_completo += ' ' + cliente.apellido_materno
        nombre_completo = nombre_completo.strip() or '-'

        row = [
            cliente.sector or '',
            getattr(cliente, 'ruta', '') or '',
            cliente.numero_libro or '',
            getattr(contrato, 'numero_contrato', '') if contrato else '',
            cliente.rut or '',
            nombre_completo,
            cliente.apellido_paterno or '',
            cliente.apellido_materno or '',
            cliente.fecha_incorporacion.strftime('%d/%m/%Y') if cliente.fecha_incorporacion else '',
            contrato.fecha_contrato.strftime('%d/%m/%Y') if contrato and contrato.fecha_contrato else '',
            cliente.sexo or '',
            contrato.direccion_arranque if contrato else '',
            cliente.direccion or '',
            contrato.servicio if contrato else '',
            'Sí' if (contrato and contrato.socio) else 'No'
        ]
        data_rows.append(row)

    titulo = "LISTADO DE SOCIOS"
    filename = f"socios_{alias}_{fecha_archivo}.pdf"
    return _build_pdf(titulo, empresa, usuario_nombre, fecha_hora, mes_anio,
                      headers, data_rows, filename)