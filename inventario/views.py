from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from empresas.models import Empresa
from inventario.models import ItemInventario
from empresas.decorators import permiso_requerido   # ✅ 导入正确

@login_required
@permiso_requerido('inventario')
def inventario_view(request, slug):
    empresa = get_object_or_404(Empresa, slug=slug)

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        cantidad = request.POST.get('cantidad')
        categoria = request.POST.get('categoria')

        if nombre and cantidad:
            ItemInventario.objects.create(
                empresa=empresa,
                nombre=nombre,
                cantidad=int(cantidad),
                categoria=categoria
            )
            return redirect('inventario', slug=slug)

    items = ItemInventario.objects.filter(empresa=empresa).order_by('-fecha_creacion')

    return render(request, 'inventario/inventario.html', {
        'empresa': empresa,
        'items': items,
        'slug': slug
    })

def agregar_item(request, slug):
    empresa = get_object_or_404(Empresa, slug=slug)

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        cantidad = request.POST.get('cantidad')
        categoria = request.POST.get('categoria')

        if nombre and cantidad:
            ItemInventario.objects.create(
                empresa=empresa,
                nombre=nombre,
                cantidad=int(cantidad),
                categoria=categoria
            )
        return redirect('inventario', slug=slug)

    return redirect('inventario', slug=slug)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from empresas.models import Empresa
from inventario.models import ItemInventario
from datetime import datetime
import pytz

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from empresas.models import Empresa
from clientes.models import Cliente
from datetime import datetime
import pytz

def escribir_cabecera_estandar(ws, empresa, usuario_nombre, fecha_hora, titulo_reporte, mes_anio=None):
    """
    Escribe la cabecera estándar en las primeras filas de la hoja Excel.
    Retorna el número de la siguiente fila disponible para comenzar a escribir datos.
    """
    row = 1
    font_normal = Font(name='Calibri', size=11)
    font_bold = Font(name='Calibri', size=11, bold=True)
    font_titulo = Font(name='Calibri', size=12, bold=True)
    alignment_left = Alignment(horizontal='left', vertical='center')

    # Fila 1: Nombre comité (mayúsculas, negrita)
    cell = ws.cell(row=row, column=1, value=empresa.nombre.upper())
    cell.font = font_bold
    cell.alignment = alignment_left
    row += 1

    # Fila 2: RUT (si el modelo Empresa tiene campo rut, si no se omite o se pone fijo)
    rut_empresa = getattr(empresa, 'rut', '') or ''
    if rut_empresa:
        cell = ws.cell(row=row, column=1, value=rut_empresa)
        cell.font = font_normal
        cell.alignment = alignment_left
        row += 1

    # Fila 3: Usuario
    cell = ws.cell(row=row, column=1, value=f"Usuario: {usuario_nombre}")
    cell.font = font_normal
    cell.alignment = alignment_left
    row += 1

    # Fila 4: Fecha y hora
    cell = ws.cell(row=row, column=1, value=fecha_hora)
    cell.font = font_normal
    cell.alignment = alignment_left
    row += 1

    # Fila 5: Título del reporte (mayúsculas, negrita)
    cell = ws.cell(row=row, column=1, value=titulo_reporte.upper())
    cell.font = font_titulo
    cell.alignment = alignment_left
    row += 1

    # Fila 6: Mes y año (opcional)
    if mes_anio:
        cell = ws.cell(row=row, column=1, value=mes_anio.upper())
        cell.font = font_normal
        cell.alignment = alignment_left
        row += 1

    row += 2  # Dejar una fila en blanco antes de los datos

    return row

def reporte_inventario_excel(request, slug):
    """
    Genera un reporte Excel con el inventario actual de la empresa.
    """
    empresa = get_object_or_404(Empresa, slug=slug)

    user = request.user
    usuario_nombre = user.get_full_name() or user.username if user.is_authenticated else "Anónimo"
    santiago_tz = pytz.timezone('America/Santiago')
    ahora = datetime.now(santiago_tz)
    fecha_hora = ahora.strftime('%d/%m/%Y %H:%M:%S')
    fecha_archivo = ahora.strftime('%Y%m%d_%H%M%S')

    meses_es = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
    mes_anio = f"{meses_es[ahora.month-1]} de {ahora.year}"

    items = ItemInventario.objects.filter(empresa=empresa).order_by('categoria', 'nombre')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    ws.sheet_view.showGridLines = False

    titulo_reporte = "INVENTARIO DE MATERIALES Y EQUIPOS"
    fila_inicio_tabla = escribir_cabecera_estandar(
        ws, empresa, usuario_nombre, fecha_hora, titulo_reporte, mes_anio
    )

    # Estilos de tabla
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    data_font = Font(name='Calibri', size=9)
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_alignment_center = Alignment(horizontal='center', vertical='center')
    data_alignment_right = Alignment(horizontal='right', vertical='center')

    full_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    headers = ['N°', 'Categoría', 'Nombre del ítem', 'Cantidad', 'Fecha de registro']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=fila_inicio_tabla, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = full_border

    if items.exists():
        for row_idx, item in enumerate(items, start=fila_inicio_tabla + 1):
            row_data = [
                row_idx - fila_inicio_tabla,  # numeración correlativa
                item.categoria,
                item.nombre,
                item.cantidad,
                item.fecha_creacion.strftime('%d/%m/%Y %H:%M') if item.fecha_creacion else ''
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                if col_idx == 4:  # cantidad alineada a la derecha
                    cell.alignment = data_alignment_right
                elif col_idx == 1:  # número correlativo centrado
                    cell.alignment = data_alignment_center
                else:
                    cell.alignment = data_alignment
                cell.border = full_border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
                if col_idx == 4 and isinstance(val, (int, float)):
                    cell.number_format = '#,##0'

        # Ajustar ancho de columnas
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            for row in range(fila_inicio_tabla, ws.max_row + 1):
                val = ws.cell(row, col_idx).value
                if val:
                    max_length = max(max_length, len(str(val)))
            if col_idx == 2:  # categoría
                adjusted = min(max_length + 3, 25)
            elif col_idx == 3:  # nombre
                adjusted = min(max_length + 3, 35)
            elif col_idx == 4:  # cantidad
                adjusted = 12
            elif col_idx == 5:  # fecha
                adjusted = 18
            else:
                adjusted = min(max_length + 3, 20)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

        ws.freeze_panes = f'A{fila_inicio_tabla + 1}'
    else:
        ws.merge_cells(f'A{fila_inicio_tabla}:E{fila_inicio_tabla}')
        cell = ws.cell(fila_inicio_tabla, 1, f'No hay ítems en el inventario de {empresa.nombre}.')
        cell.font = Font(size=12, bold=True, color='FF0000')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[fila_inicio_tabla].height = 30

    # Pie de página
    last_row = ws.max_row + 1
    ws[f'A{last_row}'] = f"Reporte generado el {ahora.strftime('%d/%m/%Y %H:%M:%S')}"
    ws[f'A{last_row}'].font = Font(name='Calibri', size=8, italic=True, color='7F8C8D')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"inventario_{slug}_{fecha_archivo}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response