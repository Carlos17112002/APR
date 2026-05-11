from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.db import transaction, DatabaseError
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count, Avg
from django.utils.timezone import make_aware
from django.views.decorators.http import require_http_methods
from django.core.exceptions import FieldError
from django.contrib.auth.decorators import login_required

from boletas.models import Boleta
from .models import LecturaMovil
from empresas.models import Empresa
from clientes.models import Cliente
import json
from decimal import Decimal, InvalidOperation
import uuid
from datetime import datetime, timedelta
import calendar
import logging
from django.db import connection

# Configurar logger
logger = logging.getLogger(__name__)

# ========== VISTAS PARA WEB ==========

from django.db.models import Q
from django.db.models.functions import ExtractYear
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from empresas.models import Empresa
from lecturas.models import LecturaMovil
from clientes.models import Cliente

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Q
from django.db.models.functions import ExtractYear
from empresas.models import Empresa
from lecturas.models import LecturaMovil
from clientes.models import Cliente
from empresas.decorators import permiso_requerido

@login_required
@permiso_requerido('lecturas')
def listado_lecturas_app(request, alias):
    """
    Vista para listar lecturas de la app móvil
    """
    empresa_obj = get_object_or_404(Empresa, slug=alias)
    
    # Verificar que el usuario tiene acceso a esta empresa (implementar según tu lógica)

    # --- Obtener parámetros GET con valores por defecto ---
    mes_param = request.GET.get('mes')
    anio_param = request.GET.get('anio')
    estado_param = request.GET.get('estado', '')
    usuario_param = request.GET.get('usuario', '')

    # Si no hay parámetro 'mes', usamos el mes actual
    if mes_param is None:
        mes_default = str(timezone.now().month)
    else:
        mes_default = mes_param

    # Si no hay parámetro 'anio', usamos el año actual (solo si el mes no es 'all')
    if anio_param is None:
        # Si el mes es 'all', no forzamos año (así no limitamos la consulta)
        if mes_default == 'all':
            anio_default = 'all'
        else:
            anio_default = str(timezone.now().year)
    else:
        anio_default = anio_param

    filtros = {
        'mes': mes_default,
        'anio': anio_default,
        'estado': estado_param,
        'usuario': usuario_param,
    }
    
    # --- Construir query usando empresa_slug ---
    query = Q(empresa_slug=alias)
    
    if filtros['mes'] != 'all':
        try:
            mes = int(filtros['mes'])
            query &= Q(fecha_lectura__month=mes)
        except:
            pass
    
    if filtros['anio'] != 'all':
        try:
            anio = int(filtros['anio'])
            query &= Q(fecha_lectura__year=anio)
        except:
            pass
    
    if filtros['estado']:
        query &= Q(estado=filtros['estado'])
    
    if filtros['usuario']:
        query &= Q(usuario_app__icontains=filtros['usuario'])
    
    # --- Obtener lecturas ---
    lecturas = LecturaMovil.objects.filter(query).order_by('-fecha_lectura', '-fecha_sincronizacion')
    
    # --- Estadísticas rápidas ---
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    
    lecturas_hoy = LecturaMovil.objects.filter(
        empresa_slug=alias,
        fecha_sincronizacion__date=hoy
    ).count()
    
    lecturas_mes = LecturaMovil.objects.filter(
        empresa_slug=alias,
        fecha_sincronizacion__date__gte=inicio_mes
    ).count()
    
    # --- Usuarios únicos ---
    usuarios = LecturaMovil.objects.filter(
        empresa_slug=alias
    ).values_list('usuario_app', flat=True).distinct()
    
    # --- Estados únicos para filtro ---
    estados_filtro = LecturaMovil.objects.filter(
        empresa_slug=alias
    ).values_list('estado', flat=True).distinct()
    
    # --- Información de clientes (desde la BD de la empresa) ---
    cliente_info = {}
    try:
        db_alias = f'db_{alias}'
        clientes = Cliente.objects.using(db_alias).filter(empresa_slug=alias).values('id', 'nombre', 'rut', 'medidor')
        for c in clientes:
            cliente_id = str(c['id'])
            cliente_info[cliente_id] = {
                'nombre': c['nombre'] or f"Cliente {c['id']}",
                'rut': c['rut'] or "No especificado",
                'medidor': c['medidor'] or "No especificado"
            }
    except Exception as e:
        print(f"Error obteniendo info de clientes: {e}")
        cliente_info = {}
    
    # --- Añadir info de cliente a cada lectura ---
    lecturas_con_info = []
    for lectura in lecturas:
        cliente_valor = str(lectura.cliente)
        
        if cliente_valor in cliente_info:
            info_cliente = cliente_info[cliente_valor]
        else:
            info_cliente = {
                'nombre': f"Cliente ID: {lectura.cliente}",
                'rut': "Información no disponible",
                'medidor': "No disponible"
            }
        
        lecturas_con_info.append({
            'lectura': lectura,
            'cliente_info': info_cliente
        })
    
    # --- Opciones para filtros ---
    meses = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]
    
    # Obtener años disponibles
    try:
        anios = LecturaMovil.objects.filter(
            empresa_slug=alias
        ).annotate(
            year=ExtractYear('fecha_lectura')
        ).values_list('year', flat=True).distinct().order_by('-year')
        
        anios = list(anios)
        
        if not anios:
            anios = [timezone.now().year]
    except Exception as e:
        print(f"Error obteniendo años: {e}")
        current_year = timezone.now().year
        anios = list(range(current_year - 4, current_year + 1))
    
    # --- Contexto para el template ---
    context = {
        'empresa': empresa_obj,
        'slug': alias,
        'lecturas_con_info': lecturas_con_info,
        'lecturas_hoy': lecturas_hoy,
        'lecturas_mes': lecturas_mes,
        'usuarios': list(usuarios),
        'estados_filtro': estados_filtro,
        'filtros': filtros,
        'meses': meses,
        'anios': anios,
        'page_title': 'Lecturas App Móvil',
    }
    
    return render(request, 'lecturas/listado_lecturas.html', context)

@login_required
def generar_boletas_lote(request, alias):
    """
    Vista para generar boletas en lote a partir de lecturas
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        empresa = get_object_or_404(Empresa, slug=alias)
        alias_db = f'db_{alias}'
        
        # Obtener parámetros
        mes = request.POST.get('mes', timezone.now().month)
        anio = request.POST.get('anio', timezone.now().year)
        sector = request.POST.get('sector', '')
        
        # Validar que no hay boletas ya generadas para este período
        boletas_existentes = Boleta.objects.using(alias_db).filter(
            mes=mes,
            anio=anio,
            empresa=empresa  # Boleta tiene FK a Empresa
        ).count()
        
        if boletas_existentes > 0:
            return JsonResponse({
                'success': False,
                'error': f'Ya existen {boletas_existentes} boletas generadas para {mes}/{anio}'
            })
        
        # Obtener lecturas completadas para el período
        # CORRECCIÓN: usar empresa_id en lugar de empresa
        query = Q(empresa_id=empresa.id, mes=mes, anio=anio, estado='completada')
        if sector and sector != 'all':
            # sector en Cliente es CharField, no relación
            query &= Q(cliente__sector=sector)
        
        lecturas = LecturaMovil.objects.using(alias_db).filter(query).select_related('cliente')
        
        if not lecturas:
            return JsonResponse({
                'success': False,
                'error': f'No hay lecturas completadas para {mes}/{anio}'
            })
        
        boletas_generadas = 0
        boletas_con_error = []
        
        with transaction.atomic(using=alias_db):
            for lectura in lecturas:
                try:
                    # Verificar si ya existe boleta para este cliente en el período
                    boleta_existente = Boleta.objects.using(alias_db).filter(
                        cliente=lectura.cliente,
                        mes=mes,
                        anio=anio
                    ).exists()
                    
                    if boleta_existente:
                        boletas_con_error.append({
                            'cliente': lectura.cliente.nombre,
                            'error': 'Ya existe boleta para este período'
                        })
                        continue
                    
                    # Calcular consumo (buscar lectura anterior)
                    lectura_anterior = LecturaMovil.objects.using(alias_db).filter(
                        cliente=lectura.cliente,
                        estado='completada'
                    ).exclude(id=lectura.id).order_by('-fecha_lectura').first()
                    
                    consumo = lectura.lectura - (lectura_anterior.lectura if lectura_anterior else Decimal('0'))
                    
                    if consumo < 0:
                        consumo = Decimal('0')
                    
                    # Calcular monto (esto es un ejemplo, ajusta según tu lógica)
                    tarifa_base = Decimal('1500')  # Ejemplo
                    valor_m3 = Decimal('850')  # Ejemplo
                    monto_total = tarifa_base + (consumo * valor_m3)
                    
                    # Crear boleta
                    boleta = Boleta.objects.using(alias_db).create(
                        empresa=empresa,
                        cliente=lectura.cliente,
                        lectura=lectura,
                        mes=mes,
                        anio=anio,
                        consumo=consumo,
                        monto_total=monto_total,
                        tarifa_base=tarifa_base,
                        valor_m3=valor_m3,
                        fecha_emision=timezone.now().date(),
                        fecha_vencimiento=timezone.now().date() + timedelta(days=15),
                        estado='pendiente',
                        numero_boleta=f"B{anio}{mes:02d}{lectura.cliente.id:06d}"
                    )
                    
                    boletas_generadas += 1
                    
                except Exception as e:
                    logger.error(f"Error generando boleta para cliente {lectura.cliente.id}: {str(e)}")
                    boletas_con_error.append({
                        'cliente': lectura.cliente.nombre,
                        'error': str(e)
                    })
        
        logger.info(f"Generadas {boletas_generadas} boletas para empresa {empresa.nombre} ({mes}/{anio})")
        
        return JsonResponse({
            'success': True,
            'empresa': empresa.nombre,
            'mes': mes,
            'anio': anio,
            'boletas_generadas': boletas_generadas,
            'boletas_con_error': len(boletas_con_error),
            'detalle_errores': boletas_con_error,
            'mensaje': f'Se generaron {boletas_generadas} boletas exitosamente'
        })
        
    except Exception as e:
        logger.error(f"Error en generar_boletas_lote: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def listado_boletas(request, alias):
    """
    Vista para listar boletas generadas
    """
    empresa = get_object_or_404(Empresa, slug=alias)
    alias_db = f'db_{alias}'
    
    # Obtener boletas
    boletas = Boleta.objects.using(alias_db).select_related(
        'cliente', 'lectura'
    ).filter(empresa=empresa).order_by('-fecha_emision')
    
    # Aplicar filtros
    estado = request.GET.get('estado')
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    sector = request.GET.get('sector')
    cliente = request.GET.get('cliente', '').strip()
    
    if estado and estado != 'all':
        boletas = boletas.filter(estado=estado)
    if mes and mes != 'all':
        boletas = boletas.filter(mes=mes)
    if anio and anio != 'all':
        boletas = boletas.filter(anio=anio)
    if sector and sector != 'all':
        # sector en Cliente es CharField, no relación
        boletas = boletas.filter(cliente__sector=sector)
    if cliente:
        boletas = boletas.filter(Q(cliente__nombre__icontains=cliente) | Q(cliente__rut__icontains=cliente))
    
    # Paginación
    paginator = Paginator(boletas, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    hoy = timezone.now().date()
    boletas_mes = boletas.filter(
        fecha_emision__month=hoy.month,
        fecha_emision__year=hoy.year
    )
    
    estadisticas = {
        'total': boletas_mes.count(),
        'pagadas': boletas_mes.filter(estado='pagada').count(),
        'pendientes': boletas_mes.filter(estado='pendiente').count(),
        'vencidas': boletas_mes.filter(estado='vencida').count(),
        'monto_total': float(boletas_mes.aggregate(Sum('monto_total'))['monto_total__sum'] or 0),
        'monto_pagado': float(boletas_mes.filter(estado='pagada').aggregate(Sum('monto_total'))['monto_total__sum'] or 0),
    }
    
    # Obtener sectores disponibles desde Cliente en la BD de la empresa
    sectores = []
    try:
        sectores = Cliente.objects.using(alias_db).filter(
            empresa_slug=alias  # si Cliente tiene este campo
        ).values_list('sector', flat=True).distinct().order_by('sector')
    except Exception as e:
        logger.error(f"Error obteniendo sectores: {e}")
    
    # Obtener años disponibles
    anios_disponibles = Boleta.objects.using(alias_db).filter(
        empresa=empresa
    ).values_list('anio', flat=True).distinct().order_by('-anio')
    
    context = {
        'empresa': empresa,
        'slug': alias,
        'page_obj': page_obj,
        'boletas': page_obj.object_list,
        'estados': Boleta.ESTADOS_BOLETA,
        'sectores': sectores,
        'anios': anios_disponibles,
        'meses': [
            (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
            (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
            (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
        ],
        'estadisticas': estadisticas,
        'filtros': {
            'estado': estado or 'all',
            'mes': mes or 'all',
            'anio': anio or 'all',
            'sector': sector or 'all',
            'cliente': cliente,
        },
        'hoy': hoy,
    }
    
    return render(request, 'lecturas/listado_boletas.html', context)

@login_required
def detalle_lectura(request, alias, lectura_id):
    """
    Vista para ver el detalle de una lectura específica
    """
    empresa = get_object_or_404(Empresa, slug=alias)
    alias_db = f'db_{alias}'
    
    # Lectura está en BD principal
    lectura = get_object_or_404(LecturaMovil.objects.select_related(
        'empresa'  # esto asume que LecturaMovil tiene FK 'empresa' al modelo Empresa
    ), id=lectura_id, empresa_id=empresa.id)  # filtrar por empresa_id
    
    # Obtener información del cliente desde la BD específica
    cliente_info = None
    try:
        with connection.cursor(using=alias_db) as cursor:
            cursor.execute("""
                SELECT nombre, rut, direccion, medidor 
                FROM clientes_cliente 
                WHERE id = %s AND empresa_slug = %s
            """, [lectura.cliente, alias])
            
            row = cursor.fetchone()
            if row:
                cliente_info = {
                    'nombre': row[0] or f"Cliente {lectura.cliente}",
                    'rut': row[1] or "No especificado",
                    'direccion': row[2] or "No especificada",
                    'medidor': row[3] or "No especificado"
                }
    except Exception as e:
        logger.error(f"Error obteniendo info del cliente: {e}")
    
    # Obtener lecturas anteriores del mismo cliente (desde BD principal)
    lecturas_anteriores = LecturaMovil.objects.filter(
        empresa_id=empresa.id,  # usar empresa_id
        cliente=lectura.cliente
    ).exclude(id=lectura_id).order_by('-fecha_lectura')[:5]
    
    # Obtener boleta relacionada si existe
    boleta_relacionada = None
    try:
        boleta_relacionada = Boleta.objects.using(alias_db).filter(
            lectura=lectura
        ).first()
    except:
        pass
    
    context = {
        'empresa': empresa,
        'slug': alias,
        'lectura': lectura,
        'cliente_info': cliente_info,
        'lecturas_anteriores': lecturas_anteriores,
        'boleta_relacionada': boleta_relacionada,
    }
    
    return render(request, 'lecturas/detalle_lectura.html', context)

@login_required
def estadisticas_lecturas(request, alias):
    """
    Vista para mostrar estadísticas de lecturas
    """
    empresa = get_object_or_404(Empresa, slug=alias)
    
    # Parámetros de tiempo
    mes = request.GET.get('mes', timezone.now().month)
    anio = request.GET.get('anio', timezone.now().year)
    periodo = request.GET.get('periodo', 'mensual')
    
    hoy = timezone.now().date()
    
    if periodo == 'mensual':
        fecha_inicio = datetime(int(anio), int(mes), 1).date()
        if int(mes) == 12:
            fecha_fin = datetime(int(anio) + 1, 1, 1).date() - timedelta(days=1)
        else:
            fecha_fin = datetime(int(anio), int(mes) + 1, 1).date() - timedelta(days=1)
    elif periodo == 'trimestral':
        trimestre = (int(mes) - 1) // 3 + 1
        mes_inicio = (trimestre - 1) * 3 + 1
        fecha_inicio = datetime(int(anio), mes_inicio, 1).date()
        fecha_fin = datetime(int(anio), mes_inicio + 3, 1).date() - timedelta(days=1)
    else:  # anual
        fecha_inicio = datetime(int(anio), 1, 1).date()
        fecha_fin = datetime(int(anio), 12, 31).date()
    
    # Consultar datos (desde BD principal)
    lecturas_periodo = LecturaMovil.objects.filter(
        empresa_id=empresa.id,  # usar empresa_id
        fecha_lectura__date__range=[fecha_inicio, fecha_fin],
        estado='cargada'
    )
    
    total_lecturas = lecturas_periodo.count()
    lecturas_por_usuario = lecturas_periodo.values('usuario_app').annotate(
        total=Count('id')
    ).order_by('-total')
    
    # Obtener información de clientes por sector (desde BD de empresa)
    lecturas_por_sector = []
    alias_db = f'db_{alias}'
    try:
        with connection.cursor(using=alias_db) as cursor:
            cursor.execute("""
                SELECT cc.sector, COUNT(lm.id) as total
                FROM lecturas_lecturamovil lm
                INNER JOIN clientes_cliente cc ON lm.cliente = cc.id
                WHERE lm.empresa_slug = %s 
                AND lm.fecha_lectura BETWEEN %s AND %s
                AND lm.estado = 'cargada'
                GROUP BY cc.sector
                ORDER BY total DESC
            """, [alias, fecha_inicio, fecha_fin])
            
            for row in cursor.fetchall():
                lecturas_por_sector.append({
                    'sector': row[0] or 'Sin sector',
                    'total': row[1],
                    'porcentaje': (row[1] * 100.0 / total_lecturas) if total_lecturas > 0 else 0
                })
    except Exception as e:
        logger.error(f"Error obteniendo lecturas por sector: {e}")
    
    lecturas_por_dia = lecturas_periodo.values('fecha_lectura__date').annotate(
        total=Count('id')
    ).order_by('fecha_lectura__date')
    
    promedio_diario = total_lecturas / ((fecha_fin - fecha_inicio).days + 1) if (fecha_fin - fecha_inicio).days > 0 else 0
    promedio_por_usuario = total_lecturas / lecturas_por_usuario.count() if lecturas_por_usuario.count() > 0 else 0
    
    # Datos para gráficos
    dias = []
    lecturas_por_dia_list = []
    for item in lecturas_por_dia:
        dias.append(item['fecha_lectura__date'].strftime('%d/%m'))
        lecturas_por_dia_list.append(item['total'])
    
    sectores = []
    lecturas_por_sector_list = []
    for item in lecturas_por_sector:
        sectores.append(item['sector'])
        lecturas_por_sector_list.append(item['total'])
    
    usuarios = []
    lecturas_por_usuario_list = []
    for item in lecturas_por_usuario:
        if item['usuario_app']:
            usuarios.append(item['usuario_app'])
            lecturas_por_usuario_list.append(item['total'])
    
    context = {
        'empresa': empresa,
        'slug': alias,
        'periodo': periodo,
        'mes': int(mes) if mes else hoy.month,
        'anio': int(anio) if anio else hoy.year,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'estadisticas': {
            'total_lecturas': total_lecturas,
            'promedio_diario': round(promedio_diario, 1),
            'promedio_por_usuario': round(promedio_por_usuario, 1),
            'usuarios_activos': lecturas_por_usuario.count(),
            'sectores_cubiertos': len(lecturas_por_sector),
        },
        'lecturas_por_usuario': list(lecturas_por_usuario),
        'lecturas_por_sector': lecturas_por_sector,
        'lecturas_por_dia': list(lecturas_por_dia),
        'datos_graficos': {
            'dias': json.dumps(dias),
            'lecturas_por_dia': json.dumps(lecturas_por_dia_list),
            'sectores': json.dumps(sectores),
            'lecturas_por_sector': json.dumps(lecturas_por_sector_list),
            'usuarios': json.dumps(usuarios),
            'lecturas_por_usuario': json.dumps(lecturas_por_usuario_list),
        },
        'meses': [
            (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
            (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
            (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
        ],
        'anios': range(hoy.year - 5, hoy.year + 1),
        'periodos': [('mensual', 'Mensual'), ('trimestral', 'Trimestral'), ('anual', 'Anual')],
    }
    
    return render(request, 'lecturas/estadisticas_lecturas.html', context)

from django.conf import settings

@login_required
def mapa_lecturas(request, alias):
    """
    Vista para mostrar en un mapa las lecturas de la app móvil,
    filtradas por mes, año y estado.
    """
    empresa = get_object_or_404(Empresa, slug=alias)

    # Obtener filtros de la request
    mes = request.GET.get('mes', timezone.now().month)
    anio = request.GET.get('anio', timezone.now().year)
    estado = request.GET.get('estado', 'cargada')

    # Convertir a enteros si es posible
    try:
        mes = int(mes)
    except:
        mes = timezone.now().month
    try:
        anio = int(anio)
    except:
        anio = timezone.now().year

    # Consulta base de lecturas con coordenadas
    lecturas_qs = LecturaMovil.objects.filter(
        empresa_id=empresa.id,
        fecha_lectura__month=mes,
        fecha_lectura__year=anio,
        latitud__isnull=False,
        longitud__isnull=False
    )
    if estado and estado != 'all':
        lecturas_qs = lecturas_qs.filter(estado=estado)

    # Construir lista de puntos para el mapa
    puntos_mapa = []
    for lectura in lecturas_qs:
        puntos_mapa.append({
            'id': str(lectura.id),
            'nombre': f"Cliente {lectura.cliente}",
            'lat': float(lectura.latitud),
            'lng': float(lectura.longitud),
            'lectura': float(lectura.lectura_actual),
            'fecha': lectura.fecha_lectura.strftime('%d/%m/%Y'),
            'usuario': lectura.usuario_app,
            'estado': lectura.estado,
            'color': {
                'cargada': 'green',
                'pendiente': 'orange',
                'procesada': 'blue'
            }.get(lectura.estado, 'gray')
        })

    # Obtener estadísticas por sector (usando la BD de la empresa)
    lecturas_por_sector = []
    alias_db = f'db_{alias}'
    if alias_db in settings.DATABASES:
        try:
            with connections[alias_db].cursor() as cursor:
                cursor.execute("""
                    SELECT cc.sector, COUNT(lm.id) as total
                    FROM lecturas_lecturamovil lm
                    INNER JOIN clientes_cliente cc ON lm.cliente = cc.id
                    WHERE lm.empresa_slug = %s 
                      AND strftime('%%m', lm.fecha_lectura) = %s
                      AND strftime('%%Y', lm.fecha_lectura) = %s
                      AND lm.estado = %s
                    GROUP BY cc.sector
                    ORDER BY total DESC
                """, [alias, f"{mes:02d}", str(anio), estado])
                
                for row in cursor.fetchall():
                    lecturas_por_sector.append({
                        'sector': row[0] or 'Sin sector',
                        'total': row[1]
                    })
        except Exception as e:
            logger.error(f"Error obteniendo lecturas por sector: {e}")
    else:
        logger.warning(f"Base de datos {alias_db} no encontrada en settings")

    # Si no hay datos, mostrar un placeholder
    if not lecturas_por_sector:
        lecturas_por_sector = [{'sector': 'Sin datos', 'total': 0}]

    # Centro del mapa (primer punto o coordenadas por defecto)
    if puntos_mapa:
        center_lat = puntos_mapa[0]['lat']
        center_lng = puntos_mapa[0]['lng']
    else:
        center_lat = -33.4489
        center_lng = -70.6693

    # Opciones para los filtros
    meses_opciones = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]
    anios_opciones = range(timezone.now().year - 5, timezone.now().year + 1)
    estados_opciones = ['cargada', 'pendiente', 'procesada']

    context = {
        'empresa': empresa,
        'slug': alias,
        'puntos_mapa': json.dumps(puntos_mapa),
        'total_puntos': len(puntos_mapa),
        'lecturas_por_sector': lecturas_por_sector,
        'filtros': {
            'mes': mes,
            'anio': anio,
            'estado': estado,
        },
        'meses': meses_opciones,
        'anios': anios_opciones,
        'estados': estados_opciones,
        'center_lat': center_lat,
        'center_lng': center_lng,
    }
    return render(request, 'lecturas/mapa_lecturas.html', context)




# ========== API PARA APP MÓVIL ==========

@csrf_exempt
@require_http_methods(["POST"])
def api_sincronizar_lecturas(request, alias):
    """
    API para sincronizar lecturas desde la app móvil - CORREGIDO: acepta 'alias'
    """
    try:
        empresa = get_object_or_404(Empresa, slug=alias)
        
        # Parsear datos
        try:
            data = json.loads(request.body)
            lecturas_data = data.get('lecturas', [])
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Formato JSON inválido'
            }, status=400)
        
        if not lecturas_data:
            return JsonResponse({
                'success': False,
                'error': 'No hay lecturas para sincronizar'
            }, status=400)
        
        lecturas_procesadas = []
        lecturas_con_error = []
        
        for lectura_data in lecturas_data:
            try:
                # Validar datos básicos
                cliente_id = lectura_data.get('cliente_id')
                lectura_valor = lectura_data.get('lectura_actual')
                
                if not cliente_id or not lectura_valor:
                    lecturas_con_error.append({
                        'cliente_id': cliente_id,
                        'error': 'Datos incompletos'
                    })
                    continue
                
                # Convertir lectura a Decimal
                try:
                    lectura_decimal = Decimal(str(lectura_valor))
                except (InvalidOperation, ValueError):
                    lecturas_con_error.append({
                        'cliente_id': cliente_id,
                        'error': 'Valor de lectura inválido'
                    })
                    continue
                
                # Crear nueva lectura en BD principal
                lectura_uuid = uuid.uuid4().hex
                
                with connection.cursor() as cursor:
                    cursor.execute("""
                        INSERT INTO lecturas_lecturamovil 
                        (id, empresa_id, cliente, fecha_lectura, lectura_actual, 
                         latitud, longitud, estado, observaciones_app, usuario_app, 
                         empresa_slug, usada_para_boleta, fecha_sincronizacion)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, [
                        lectura_uuid,
                        empresa.id,
                        int(cliente_id),
                        timezone.now().date(),
                        str(lectura_decimal),
                        lectura_data.get('latitud'),
                        lectura_data.get('longitud'),
                        'cargada',
                        lectura_data.get('observaciones', ''),
                        lectura_data.get('usuario', 'App Móvil'),
                        alias,
                        0,
                        timezone.now()
                    ])
                
                lecturas_procesadas.append({
                    'cliente_id': cliente_id,
                    'lectura_id': lectura_uuid,
                    'estado': 'creada',
                    'fecha': timezone.now().isoformat()
                })
                
            except Exception as e:
                logger.error(f"Error procesando lectura para cliente {lectura_data.get('cliente_id')}: {str(e)}")
                lecturas_con_error.append({
                    'cliente_id': lectura_data.get('cliente_id'),
                    'error': str(e)
                })
        
        return JsonResponse({
            'success': True,
            'empresa': empresa.nombre,
            'lecturas_procesadas': len(lecturas_procesadas),
            'lecturas_con_error': len(lecturas_con_error),
            'detalle_procesadas': lecturas_procesadas,
            'detalle_errores': lecturas_con_error,
            'fecha_sincronizacion': timezone.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Error en api_sincronizar_lecturas: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def api_obtener_clientes_pendientes(request, alias):
    """
    API para obtener clientes con lecturas pendientes - CORREGIDO: acepta 'alias'
    """
    try:
        empresa = get_object_or_404(Empresa, slug=alias)
        
        # Parámetros de filtro
        limite = int(request.GET.get('limite', 100))
        
        # Obtener clientes desde la BD específica
        clientes_data = []
        try:
            alias_db = f'db_{alias}'
            if alias_db in connection.settings_dict['DATABASES']:
                with connection.cursor() as cursor:
                    cursor.execute(f"""
                        SELECT id, nombre, rut, direccion, medidor, sector
                        FROM clientes_cliente 
                        WHERE empresa_slug = %s AND activo = 1
                        LIMIT %s
                    """, [alias, limite])
                    
                    for row in cursor.fetchall():
                        clientes_data.append({
                            'id': row[0],
                            'nombre': row[1],
                            'rut': row[2],
                            'direccion': row[3],
                            'medidor': row[4],
                            'sector': row[5]
                        })
        except Exception as e:
            print(f"Error obteniendo clientes: {e}")
        
        return JsonResponse({
            'success': True,
            'empresa': empresa.nombre,
            'total_clientes': len(clientes_data),
            'clientes': clientes_data,
            'fecha_consulta': timezone.now().date().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Error en api_obtener_clientes_pendientes: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    

from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required

@login_required
def calcular_consumo(request, alias, lectura_id):
    """Vista para calcular consumo de una lectura"""
    db_alias = f'db_{alias}'
    lectura = get_object_or_404(LecturaMovil.objects.using(db_alias), id=lectura_id)
    
    if request.method == 'POST':
        # Calcular consumo si hay lectura anterior
        if lectura.lectura_anterior:
            lectura.calcular_consumo()
            # Mensaje de éxito
            # Redirigir al detalle
    
    return redirect('detalle_lectura', alias=alias, lectura_id=lectura_id)

# ========== API ESPECÍFICA PARA APP MÓVIL ==========

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json
from django.db import connection
from django.utils import timezone
from django.shortcuts import get_object_or_404
from .models import DispositivoMovil, ConfigAppMovil
from empresas.models import Empresa
import hashlib
import secrets

@csrf_exempt
def api_dispositivo_login(request, alias):
    """
    API para autenticar dispositivo móvil
    URL: /api/<alias>/dispositivos/login/
    """
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Método no permitido'
        }, status=405)
    
    try:
        empresa = get_object_or_404(Empresa, slug=alias)
        
        # Parsear datos
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Formato JSON inválido'
            }, status=400)
        
        dispositivo_id = data.get('dispositivo_id')
        token = data.get('token')
        
        if not dispositivo_id or not token:
            return JsonResponse({
                'success': False,
                'error': 'Dispositivo ID y token requeridos'
            }, status=400)
        
        # Buscar dispositivo
        try:
            dispositivo = DispositivoMovil.objects.get(
                identificador=dispositivo_id,
                token_acceso=token,
                empresa=empresa,
                activo=True
            )
        except DispositivoMovil.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Dispositivo no encontrado o inactivo'
            }, status=401)
        
        # Actualizar última conexión
        dispositivo.ultima_conexion = timezone.now()
        dispositivo.save()
        
        # Obtener configuración de la app
        config_app, _ = ConfigAppMovil.objects.get_or_create(empresa=empresa)
        
        return JsonResponse({
            'success': True,
            'token': dispositivo.token_acceso,
            'dispositivo_nombre': dispositivo.nombre_dispositivo,
            'dispositivo_id': dispositivo.identificador,
            'empresa': {
                'nombre': empresa.nombre,
                'slug': empresa.slug,
                'version_app': empresa.version_app,
                'color_primario': empresa.color_app_primario,
                'color_secundario': empresa.color_app_secundario,
                'url_servidor': empresa.url_servidor,
            },
            'configuracion': {
                'habilitar_mapa': config_app.habilitar_mapa,
                'habilitar_offline': config_app.habilitar_offline,
                'validar_gps': config_app.validar_gps,
                'sincronizacion_auto': config_app.sincronizacion_auto,
                'mostrar_logo': config_app.mostrar_logo,
                'intervalo_sincronizacion': config_app.intervalo_sincronizacion,
                'mensaje_bienvenida': config_app.mensaje_bienvenida,
                'max_lecturas_pendientes': config_app.max_lecturas_pendientes,
            },
            'fecha_sincronizacion': timezone.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Error en api_dispositivo_login: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
def api_obtener_sectores(request, alias):
    """
    API para obtener sectores de la empresa
    URL: /api/<alias>/sectores/
    """
    if request.method != 'GET':
        return JsonResponse({
            'success': False,
            'error': 'Método no permitido'
        }, status=405)
    
    try:
        empresa = get_object_or_404(Empresa, slug=alias)
        
        # Verificar autenticación del dispositivo
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Token '):
            return JsonResponse({
                'success': False,
                'error': 'Token de autenticación requerido'
            }, status=401)
        
        token = auth_header.split(' ')[1]
        
        try:
            dispositivo = DispositivoMovil.objects.get(
                token_acceso=token,
                empresa=empresa,
                activo=True
            )
        except DispositivoMovil.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Dispositivo no autenticado'
            }, status=401)
        
        # Obtener sectores desde la BD específica
        sectores = []
        alias_db = f'db_{alias}'
        
        if alias_db in connection.settings_dict['DATABASES']:
            with connection.cursor() as cursor:
                cursor.execute(f"""
                    SELECT DISTINCT sector 
                    FROM clientes_cliente 
                    WHERE empresa_slug = %s AND activo = 1
                    ORDER BY sector
                """, [alias])
                
                for row in cursor.fetchall():
                    if row[0]:  # Solo agregar si no es None o vacío
                        sectores.append(row[0])
        
        # Si no hay sectores en clientes, usar los de la empresa
        if not sectores:
            sectores = empresa.sectores() or []
        
        return JsonResponse({
            'success': True,
            'empresa': empresa.nombre,
            'total_sectores': len(sectores),
            'sectores': sectores,
            'fecha_consulta': timezone.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Error en api_obtener_sectores: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
def api_obtener_clientes_por_sector(request, alias, sector):
    """
    API para obtener clientes de un sector específico
    URL: /api/<alias>/sectores/<sector>/clientes/
    """
    if request.method != 'GET':
        return JsonResponse({
            'success': False,
            'error': 'Método no permitido'
        }, status=405)
    
    try:
        empresa = get_object_or_404(Empresa, slug=alias)
        
        # Verificar autenticación
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Token '):
            return JsonResponse({
                'success': False,
                'error': 'Token de autenticación requerido'
            }, status=401)
        
        token = auth_header.split(' ')[1]
        
        try:
            dispositivo = DispositivoMovil.objects.get(
                token_acceso=token,
                empresa=empresa,
                activo=True
            )
        except DispositivoMovil.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Dispositivo no autenticado'
            }, status=401)
        
        # Obtener clientes del sector
        clientes = []
        alias_db = f'db_{alias}'
        
        if alias_db in connection.settings_dict['DATABASES']:
            with connection.cursor() as cursor:
                cursor.execute(f"""
                    SELECT id, nombre, rut, direccion, medidor, sector
                    FROM clientes_cliente 
                    WHERE empresa_slug = %s AND activo = 1 AND sector = %s
                    ORDER BY nombre
                """, [alias, sector])
                
                for row in cursor.fetchall():
                    # Obtener última lectura si existe
                    cursor.execute(f"""
                        SELECT lectura_actual, fecha_lectura 
                        FROM lecturas_lecturamovil 
                        WHERE empresa_slug = %s AND cliente = %s
                        ORDER BY fecha_lectura DESC 
                        LIMIT 1
                    """, [alias, row[0]])
                    
                    ultima_lectura = cursor.fetchone()
                    
                    clientes.append({
                        'id': row[0],
                        'nombre': row[1] or f"Cliente {row[0]}",
                        'rut': row[2] or "No especificado",
                        'direccion': row[3] or "No especificada",
                        'medidor': row[4] or "No especificado",
                        'sector': row[5] or "Sin sector",
                        'ultima_lectura': {
                            'valor': float(ultima_lectura[0]) if ultima_lectura and ultima_lectura[0] else 0,
                            'fecha': ultima_lectura[1].isoformat() if ultima_lectura and ultima_lectura[1] else None
                        } if ultima_lectura else None,
                        'pendiente': not bool(ultima_lectura)  # Pendiente si no tiene lecturas
                    })
        
        return JsonResponse({
            'success': True,
            'empresa': empresa.nombre,
            'sector': sector,
            'total_clientes': len(clientes),
            'clientes': clientes,
            'fecha_consulta': timezone.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Error en api_obtener_clientes_por_sector: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
def api_guardar_lectura(request, alias):
    """
    API para guardar una lectura desde la app móvil
    URL: /api/<alias>/lecturas/guardar/
    """
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Método no permitido'
        }, status=405)
    
    try:
        empresa = get_object_or_404(Empresa, slug=alias)
        
        # Verificar autenticación
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Token '):
            return JsonResponse({
                'success': False,
                'error': 'Token de autenticación requerido'
            }, status=401)
        
        token = auth_header.split(' ')[1]
        
        try:
            dispositivo = DispositivoMovil.objects.get(
                token_acceso=token,
                empresa=empresa,
                activo=True
            )
        except DispositivoMovil.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Dispositivo no autenticado'
            }, status=401)
        
        # Parsear datos
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Formato JSON inválido'
            }, status=400)
        
        # Validar datos requeridos
        cliente_id = data.get('cliente_id')
        lectura_valor = data.get('lectura')
        latitud = data.get('latitud')
        longitud = data.get('longitud')
        
        if not cliente_id or not lectura_valor:
            return JsonResponse({
                'success': False,
                'error': 'Cliente ID y lectura requeridos'
            }, status=400)
        
        # Verificar que el cliente existe
        alias_db = f'db_{alias}'
        cliente_existe = False
        if alias_db in connection.settings_dict['DATABASES']:
            with connection.cursor() as cursor:
                cursor.execute(f"""
                    SELECT COUNT(*) 
                    FROM clientes_cliente 
                    WHERE id = %s AND empresa_slug = %s AND activo = 1
                """, [cliente_id, alias])
                
                cliente_existe = cursor.fetchone()[0] > 0
        
        if not cliente_existe:
            return JsonResponse({
                'success': False,
                'error': f'Cliente {cliente_id} no encontrado o inactivo'
            }, status=404)
        
        # Configuración de la app para validaciones
        config_app, _ = ConfigAppMovil.objects.get_or_create(empresa=empresa)
        
        # Validar GPS si está habilitado
        if config_app.validar_gps and (not latitud or not longitud):
            return JsonResponse({
                'success': False,
                'error': 'Ubicación GPS requerida'
            }, status=400)
        
        # Crear lectura en BD principal
        from decimal import Decimal, InvalidOperation
        
        try:
            lectura_decimal = Decimal(str(lectura_valor))
        except (InvalidOperation, ValueError):
            return JsonResponse({
                'success': False,
                'error': 'Valor de lectura inválido'
            }, status=400)
        
        import uuid
        from datetime import datetime
        
        lectura_uuid = str(uuid.uuid4())
        
        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO lecturas_lecturamovil 
                (id, empresa_id, cliente, fecha_lectura, lectura_actual, 
                 latitud, longitud, estado, observaciones_app, usuario_app, 
                 empresa_slug, dispositivo_id, fecha_sincronizacion)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, [
                lectura_uuid,
                empresa.id,
                int(cliente_id),
                datetime.now().date(),
                str(lectura_decimal),
                latitud,
                longitud,
                'pendiente',
                data.get('observaciones', ''),
                dispositivo.nombre_dispositivo,
                alias,
                dispositivo.identificador,
                timezone.now()
            ])
        
        # Actualizar última conexión del dispositivo
        dispositivo.ultima_conexion = timezone.now()
        dispositivo.save()
        
        return JsonResponse({
            'success': True,
            'lectura_id': lectura_uuid,
            'cliente_id': cliente_id,
            'lectura': float(lectura_decimal),
            'fecha': datetime.now().isoformat(),
            'dispositivo': dispositivo.nombre_dispositivo,
            'estado': 'pendiente',
            'mensaje': 'Lectura guardada exitosamente'
        })
        
    except Exception as e:
        logger.error(f"Error en api_guardar_lectura: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
def api_obtener_lecturas_pendientes(request, alias):
    """
    API para obtener lecturas pendientes del dispositivo
    URL: /api/<alias>/lecturas/pendientes/
    """
    if request.method != 'GET':
        return JsonResponse({
            'success': False,
            'error': 'Método no permitido'
        }, status=405)
    
    try:
        empresa = get_object_or_404(Empresa, slug=alias)
        
        # Verificar autenticación
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Token '):
            return JsonResponse({
                'success': False,
                'error': 'Token de autenticación requerido'
            }, status=401)
        
        token = auth_header.split(' ')[1]
        
        try:
            dispositivo = DispositivoMovil.objects.get(
                token_acceso=token,
                empresa=empresa,
                activo=True
            )
        except DispositivoMovil.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Dispositivo no autenticado'
            }, status=401)
        
        # Obtener lecturas pendientes del dispositivo
        lecturas_pendientes = []
        
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, cliente, lectura_actual, latitud, longitud, 
                       observaciones_app, fecha_lectura, estado
                FROM lecturas_lecturamovil 
                WHERE empresa_slug = %s AND dispositivo_id = %s AND estado = 'pendiente'
                ORDER BY fecha_lectura DESC
            """, [alias, dispositivo.identificador])
            
            for row in cursor.fetchall():
                lecturas_pendientes.append({
                    'id': row[0],
                    'cliente_id': row[1],
                    'lectura': float(row[2]) if row[2] else 0,
                    'latitud': float(row[3]) if row[3] else None,
                    'longitud': float(row[4]) if row[4] else None,
                    'observaciones': row[5] or '',
                    'fecha': row[6].isoformat() if row[6] else None,
                    'estado': row[7]
                })
        
        return JsonResponse({
            'success': True,
            'empresa': empresa.nombre,
            'dispositivo': dispositivo.nombre_dispositivo,
            'total_pendientes': len(lecturas_pendientes),
            'lecturas': lecturas_pendientes,
            'fecha_consulta': timezone.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Error en api_obtener_lecturas_pendientes: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
def api_sincronizar_lecturas_batch(request, alias):
    """
    API para sincronizar múltiples lecturas en lote
    URL: /api/<alias>/lecturas/sincronizar/
    """
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Método no permitido'
        }, status=405)
    
    try:
        empresa = get_object_or_404(Empresa, slug=alias)
        
        # Verificar autenticación
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Token '):
            return JsonResponse({
                'success': False,
                'error': 'Token de autenticación requerido'
            }, status=401)
        
        token = auth_header.split(' ')[1]
        
        try:
            dispositivo = DispositivoMovil.objects.get(
                token_acceso=token,
                empresa=empresa,
                activo=True
            )
        except DispositivoMovil.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Dispositivo no autenticado'
            }, status=401)
        
        # Parsear datos
        try:
            data = json.loads(request.body)
            lecturas_data = data.get('lecturas', [])
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Formato JSON inválido'
            }, status=400)
        
        if not lecturas_data:
            return JsonResponse({
                'success': False,
                'error': 'No hay lecturas para sincronizar'
            }, status=400)
        
        lecturas_procesadas = []
        lecturas_con_error = []
        
        for lectura_data in lecturas_data:
            try:
                # Validar datos mínimos
                if not all(k in lectura_data for k in ['cliente_id', 'lectura']):
                    lecturas_con_error.append({
                        'cliente_id': lectura_data.get('cliente_id'),
                        'error': 'Datos incompletos'
                    })
                    continue
                
                cliente_id = lectura_data['cliente_id']
                
                # Verificar que el cliente existe
                alias_db = f'db_{alias}'
                cliente_existe = False
                if alias_db in connection.settings_dict['DATABASES']:
                    with connection.cursor() as cursor:
                        cursor.execute(f"""
                            SELECT COUNT(*) 
                            FROM clientes_cliente 
                            WHERE id = %s AND empresa_slug = %s AND activo = 1
                        """, [cliente_id, alias])
                        
                        cliente_existe = cursor.fetchone()[0] > 0
                
                if not cliente_existe:
                    lecturas_con_error.append({
                        'cliente_id': cliente_id,
                        'error': 'Cliente no encontrado'
                    })
                    continue
                
                # Convertir lectura a Decimal
                from decimal import Decimal, InvalidOperation
                
                try:
                    lectura_decimal = Decimal(str(lectura_data['lectura']))
                except (InvalidOperation, ValueError):
                    lecturas_con_error.append({
                        'cliente_id': cliente_id,
                        'error': 'Valor de lectura inválido'
                    })
                    continue
                
                # Crear o actualizar lectura
                import uuid
                from datetime import datetime
                
                # Intentar encontrar lectura existente
                with connection.cursor() as cursor:
                    cursor.execute("""
                        SELECT id FROM lecturas_lecturamovil 
                        WHERE empresa_slug = %s AND cliente = %s AND dispositivo_id = %s
                        AND estado = 'pendiente'
                        LIMIT 1
                    """, [alias, cliente_id, dispositivo.identificador])
                    
                    existing = cursor.fetchone()
                    
                    if existing:
                        # Actualizar lectura existente
                        cursor.execute("""
                            UPDATE lecturas_lecturamovil 
                            SET lectura_actual = %s, latitud = %s, longitud = %s,
                                observaciones_app = %s, fecha_sincronizacion = %s,
                                estado = 'cargada'
                            WHERE id = %s
                        """, [
                            str(lectura_decimal),
                            lectura_data.get('latitud'),
                            lectura_data.get('longitud'),
                            lectura_data.get('observaciones', ''),
                            timezone.now(),
                            existing[0]
                        ])
                        
                        lectura_id = existing[0]
                    else:
                        # Crear nueva lectura
                        lectura_id = str(uuid.uuid4())
                        
                        cursor.execute("""
                            INSERT INTO lecturas_lecturamovil 
                            (id, empresa_id, cliente, fecha_lectura, lectura_actual, 
                             latitud, longitud, estado, observaciones_app, usuario_app, 
                             empresa_slug, dispositivo_id, fecha_sincronizacion)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, [
                            lectura_id,
                            empresa.id,
                            int(cliente_id),
                            datetime.now().date(),
                            str(lectura_decimal),
                            lectura_data.get('latitud'),
                            lectura_data.get('longitud'),
                            'cargada',
                            lectura_data.get('observaciones', ''),
                            dispositivo.nombre_dispositivo,
                            alias,
                            dispositivo.identificador,
                            timezone.now()
                        ])
                
                lecturas_procesadas.append({
                    'cliente_id': cliente_id,
                    'lectura_id': lectura_id,
                    'estado': 'sincronizada',
                    'fecha': timezone.now().isoformat()
                })
                
            except Exception as e:
                logger.error(f"Error procesando lectura para cliente {lectura_data.get('cliente_id')}: {str(e)}")
                lecturas_con_error.append({
                    'cliente_id': lectura_data.get('cliente_id'),
                    'error': str(e)
                })
        
        # Actualizar última conexión del dispositivo
        dispositivo.ultima_conexion = timezone.now()
        dispositivo.save()
        
        return JsonResponse({
            'success': True,
            'empresa': empresa.nombre,
            'dispositivo': dispositivo.nombre_dispositivo,
            'lecturas_procesadas': len(lecturas_procesadas),
            'lecturas_con_error': len(lecturas_con_error),
            'detalle_procesadas': lecturas_procesadas,
            'detalle_errores': lecturas_con_error,
            'fecha_sincronizacion': timezone.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Error en api_sincronizar_lecturas_batch: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
def api_validar_gps(request, alias):
    """
    API para validar coordenadas GPS
    URL: /api/<alias>/validar-gps/
    """
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Método no permitido'
        }, status=405)
    
    try:
        empresa = get_object_or_404(Empresa, slug=alias)
        
        # Verificar autenticación
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Token '):
            return JsonResponse({
                'success': False,
                'error': 'Token de autenticación requerido'
            }, status=401)
        
        token = auth_header.split(' ')[1]
        
        try:
            dispositivo = DispositivoMovil.objects.get(
                token_acceso=token,
                empresa=empresa,
                activo=True
            )
        except DispositivoMovil.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Dispositivo no autenticado'
            }, status=401)
        
        # Parsear datos
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Formato JSON inválido'
            }, status=400)
        
        latitud = data.get('latitud')
        longitud = data.get('longitud')
        cliente_id = data.get('cliente_id')
        
        if not latitud or not longitud:
            return JsonResponse({
                'success': False,
                'error': 'Coordenadas GPS requeridas'
            }, status=400)
        
        # Validar que las coordenadas sean números válidos
        try:
            lat = float(latitud)
            lng = float(longitud)
        except ValueError:
            return JsonResponse({
                'success': False,
                'error': 'Coordenadas inválidas'
            }, status=400)
        
        # Validar rangos de coordenadas (Chile aproximadamente)
        # Latitud: -56 a -17, Longitud: -76 a -66
        if not (-56 <= lat <= -17) or not (-76 <= lng <= -66):
            return JsonResponse({
                'success': False,
                'error': 'Ubicación fuera de rango válido',
                'detalle': 'Las coordenadas deben estar dentro de Chile'
            })
        
        # Si se proporciona cliente_id, validar que esté cerca del cliente
        if cliente_id:
            alias_db = f'db_{alias}'
            if alias_db in connection.settings_dict['DATABASES']:
                with connection.cursor() as cursor:
                    cursor.execute(f"""
                        SELECT direccion, coordenadas 
                        FROM clientes_cliente 
                        WHERE id = %s AND empresa_slug = %s
                    """, [cliente_id, alias])
                    
                    row = cursor.fetchone()
                    if row and row[1]:  # Si tiene coordenadas guardadas
                        # Aquí podrías implementar validación de proximidad
                        # Por ahora solo registramos la validación
                        pass
        
        return JsonResponse({
            'success': True,
            'validado': True,
            'latitud': lat,
            'longitud': lng,
            'cliente_id': cliente_id,
            'mensaje': 'Ubicación GPS válida',
            'fecha_validacion': timezone.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Error en api_validar_gps: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@csrf_exempt
def api_descargar_config_app(request, alias):
    """
    API para descargar configuración de la app móvil
    URL: /api/<alias>/descargar-app/
    """
    if request.method != 'GET':
        return JsonResponse({
            'success': False,
            'error': 'Método no permitido'
        }, status=405)
    
    try:
        empresa = get_object_or_404(Empresa, slug=alias)
        config_app, _ = ConfigAppMovil.objects.get_or_create(empresa=empresa)
        
        # Construir configuración completa
        config_completa = {
            'app_name': f'SSR {empresa.nombre}',
            'empresa_slug': empresa.slug,
            'version': empresa.version_app,
            'primary_color': empresa.color_app_primario or '#1E40AF',
            'secondary_color': empresa.color_app_secundario or '#DC2626',
            'base_url': f'{empresa.url_servidor}/api/{empresa.slug}/',
            'sectores': empresa.sectores(),
            'habilitar_mapa': config_app.habilitar_mapa,
            'habilitar_offline': config_app.habilitar_offline,
            'validar_gps': config_app.validar_gps,
            'sincronizacion_auto': config_app.sincronizacion_auto,
            'mostrar_logo': config_app.mostrar_logo,
            'intervalo_sincronizacion': config_app.intervalo_sincronizacion,
            'mensaje_bienvenida': config_app.mensaje_bienvenida,
            'max_lecturas_pendientes': config_app.max_lecturas_pendientes,
            'fecha_generacion': timezone.now().isoformat(),
            'empresa_info': {
                'nombre': empresa.nombre,
                'descripcion': f'Aplicación móvil para lecturas de {empresa.nombre}',
                'contacto': empresa.contacto or 'Contacto no especificado',
                'telefono': empresa.telefono or 'Teléfono no especificado',
            }
        }
        
        response = JsonResponse(config_completa)
        response['Content-Disposition'] = f'attachment; filename="{empresa.slug}_config.json"'
        return response
        
    except Exception as e:
        logger.error(f"Error en api_descargar_config_app: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
from django.views.decorators.http import require_http_methods, require_POST, require_GET
        
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from decimal import Decimal
from datetime import datetime
from django.utils import timezone
import logging

from empresas.models import Empresa
from clientes.models import Cliente
from lecturas.models import LecturaMovil

logger = logging.getLogger(__name__)

@login_required
@require_POST
def registrar_lectura_ajax(request, alias, cliente_id):
    try:
        empresa_obj = get_object_or_404(Empresa, slug=alias)
        db_alias = f'db_{alias}'
        cliente = get_object_or_404(Cliente.objects.using(db_alias), id=cliente_id)

        fecha_str = request.POST.get('fecha')
        lectura_actual = request.POST.get('lectura_actual')
        observaciones = request.POST.get('observaciones', '')

        if not lectura_actual:
            return JsonResponse({'success': False, 'error': 'El valor de lectura es obligatorio.'})
        try:
            lectura_actual = Decimal(lectura_actual)
        except:
            return JsonResponse({'success': False, 'error': 'Valor de lectura inválido.'})

        # Procesar fecha
        try:
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except:
            fecha = timezone.now().date()

        # Última lectura del cliente
        ultima_lectura = LecturaMovil.objects.filter(
            empresa_slug=alias,
            cliente=cliente.id
        ).order_by('-fecha_lectura').first()

        lectura_anterior = ultima_lectura.lectura_actual if ultima_lectura else Decimal('0')
        consumo = lectura_actual - lectura_anterior
        if consumo < 0:
            consumo = Decimal('0')
            estado = 'pendiente'
        else:
            estado = 'cargada'

        # Crear nueva lectura
        nueva_lectura = LecturaMovil.objects.create(
            empresa_id=empresa_obj.id,
            cliente=cliente.id,
            fecha_lectura=fecha,
            lectura_actual=lectura_actual,
            lectura_anterior=lectura_anterior,
            consumo=consumo,
            estado=estado,
            observaciones_app=observaciones,
            usuario_app=request.user.username,
            empresa_slug=alias,
        )

        # Construir respuesta completa para la tabla
        lectura_data = {
            'id': nueva_lectura.id,
            'periodo': nueva_lectura.fecha_lectura.strftime('%m/%Y'),
            'fecha_lectura_anterior': ultima_lectura.fecha_lectura.strftime('%d/%m/%Y') if ultima_lectura else None,
            'lectura_anterior': float(lectura_anterior),
            'fecha_lectura_actual': nueva_lectura.fecha_lectura.strftime('%d/%m/%Y'),
            'lectura_actual': float(lectura_actual),
            'consumo': float(consumo),
            'cambio_medidor': False,
            'termino_medio': None,
            'saldo_promedio_anterior': None,
            'consumo_facturado': float(consumo),
            'abono_proximo_periodo': None,
            'codigo_lectura': None,
        }

        return JsonResponse({'success': True, 'lectura': lectura_data})

    except Exception as e:
        logger.error(f"Error en registrar_lectura_ajax: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)})
    
@login_required
@require_POST
def eliminar_lectura_ajax(request, alias, lectura_id):
    """
    Elimina una lectura (solo si no está asociada a una boleta).
    """
    try:
        # LecturaMovil está en la base de datos default, no en la base de la empresa
        lectura = get_object_or_404(LecturaMovil, id=lectura_id)
        
        # Opcional: verificar que la lectura pertenece a la empresa (por seguridad)
        if lectura.empresa_slug != alias:
            return JsonResponse({'success': False, 'error': 'La lectura no pertenece a esta empresa.'})
        
        # Opcional: evitar eliminar si ya tiene boleta asociada
        if lectura.usada_para_boleta:
            return JsonResponse({'success': False, 'error': 'No se puede eliminar una lectura ya usada en boleta.'})
        
        lectura.delete()
        return JsonResponse({'success': True, 'message': 'Lectura eliminada correctamente.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# lecturas/views.py

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.db.models import Sum
from django.db.models.functions import TruncMonth, TruncYear
from datetime import datetime, timedelta

from empresas.models import Empresa
from lecturas.models import LecturaMovil
from clientes.models import Cliente  # Importa el modelo Cliente


import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from empresas.models import Empresa
from clientes.models import CambioMedidor, Cliente  

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from empresas.models import Empresa
from clientes.models import Cliente
from datetime import datetime
import pytz

def escribir_cabecera_estandar(ws, empresa, usuario_nombre, fecha_hora, titulo_reporte, mes_anio=None):
    """
    Escribe la cabecera estándar en las primeras filas de la hoja Excel.
    Retorna el número de la siguiente fila disponible para comenzar a escribir datos.
    """
    row = 1
    font_normal = Font(name='Calibri', size=11)
    font_bold = Font(name='Calibri', size=11, bold=True)
    font_titulo = Font(name='Calibri', size=12, bold=True)
    alignment_left = Alignment(horizontal='left', vertical='center')

    # Fila 1: Nombre comité (mayúsculas, negrita)
    cell = ws.cell(row=row, column=1, value=empresa.nombre.upper())
    cell.font = font_bold
    cell.alignment = alignment_left
    row += 1

    # Fila 2: RUT (si el modelo Empresa tiene campo rut, si no se omite o se pone fijo)
    rut_empresa = getattr(empresa, 'rut', '') or ''
    if rut_empresa:
        cell = ws.cell(row=row, column=1, value=rut_empresa)
        cell.font = font_normal
        cell.alignment = alignment_left
        row += 1

    # Fila 3: Usuario
    cell = ws.cell(row=row, column=1, value=f"Usuario: {usuario_nombre}")
    cell.font = font_normal
    cell.alignment = alignment_left
    row += 1

    # Fila 4: Fecha y hora
    cell = ws.cell(row=row, column=1, value=fecha_hora)
    cell.font = font_normal
    cell.alignment = alignment_left
    row += 1

    # Fila 5: Título del reporte (mayúsculas, negrita)
    cell = ws.cell(row=row, column=1, value=titulo_reporte.upper())
    cell.font = font_titulo
    cell.alignment = alignment_left
    row += 1

    # Fila 6: Mes y año (opcional)
    if mes_anio:
        cell = ws.cell(row=row, column=1, value=mes_anio.upper())
        cell.font = font_normal
        cell.alignment = alignment_left
        row += 1

    row += 2  # Dejar una fila en blanco antes de los datos

    return row


def reporte_cambio_medidor(request, alias):
    db_alias = f'db_{alias}'
    empresa = get_object_or_404(Empresa, slug=alias)

    user = request.user
    usuario_nombre = user.get_full_name() or user.username if user.is_authenticated else "Anónimo"
    santiago_tz = pytz.timezone('America/Santiago')
    ahora = datetime.now(santiago_tz)
    fecha_hora = ahora.strftime('%d/%m/%Y %H:%M:%S')
    fecha_archivo = ahora.strftime('%Y%m%d_%H%M%S')

    meses_es = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
    mes_anio = f"{meses_es[ahora.month-1]} de {ahora.year}"

    cambios = CambioMedidor.objects.using(db_alias).order_by('-fecha_registro')
    print(f"[DEBUG] Cambios encontrados en {db_alias}: {cambios.count()}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cambios de Medidor"
    ws.sheet_view.showGridLines = False

    titulo_reporte = "LISTADO DE CAMBIOS DE MEDIDOR"
    fila_inicio_tabla = escribir_cabecera_estandar(
        ws, empresa, usuario_nombre, fecha_hora, titulo_reporte, mes_anio
    )

    # Estilos de tabla
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    data_font = Font(name='Calibri', size=9)
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_alignment_right = Alignment(horizontal='right', vertical='center')

    full_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    headers = [
        'Fecha Registro', 'Usuario', 'Periodo',
        'Cliente', 'RUT',
        'Medidor Retirado Marca', 'N° Retirado', 'Año Retirado',
        'Fecha Lectura Ant.', 'Lectura Ant.', 'Lectura Retiro', 'Consumo Final',
        'Medidor Nuevo Marca', 'N° Nuevo', 'Año Nuevo',
        'Fecha Instalación', 'Lectura Inicial'
    ]

    if cambios.exists():
        # Encabezados
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=fila_inicio_tabla, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = full_border

        # Datos
        for row_idx, cambio in enumerate(cambios, start=fila_inicio_tabla + 1):
            cliente = cambio.cliente
            row_data = [
                cambio.fecha_registro.strftime('%d/%m/%Y %H:%M') if cambio.fecha_registro else '',
                cambio.usuario or '',
                cambio.periodo or '',
                cliente.nombre if cliente else '',
                cliente.rut if cliente else '',
                cambio.medidor_retirado_marca or '',
                cambio.medidor_retirado_numero,
                cambio.medidor_retirado_anio or '',
                cambio.fecha_lectura_anterior.strftime('%d/%m/%Y') if cambio.fecha_lectura_anterior else '',
                float(cambio.lectura_anterior) if cambio.lectura_anterior else '',
                float(cambio.lectura_retiro),
                float(cambio.consumo_final) if cambio.consumo_final else '',
                cambio.medidor_nuevo_marca or '',
                cambio.medidor_nuevo_numero,
                cambio.medidor_nuevo_anio or '',
                cambio.fecha_instalacion.strftime('%d/%m/%Y') if cambio.fecha_instalacion else '',
                float(cambio.lectura_inicial)
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row_idx, col_idx, val)
                cell.font = data_font
                cell.alignment = data_alignment_right if isinstance(val, (int, float)) else data_alignment
                cell.border = full_border
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    cell.number_format = '#,##0.00'
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # Ajustar ancho de columnas
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            for row in range(fila_inicio_tabla, ws.max_row + 1):
                val = ws.cell(row, col_idx).value
                if val:
                    max_length = max(max_length, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 3, 45)

        ws.freeze_panes = f'A{fila_inicio_tabla + 1}'
    else:
        ws.merge_cells(f'A{fila_inicio_tabla}:Q{fila_inicio_tabla}')
        cell = ws.cell(fila_inicio_tabla, 1, f'No hay cambios de medidor registrados para {empresa.nombre}.')
        cell.font = Font(size=12, bold=True, color='FF0000')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[fila_inicio_tabla].height = 30

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="cambio_medidor_{alias}_{fecha_archivo}.xlsx"'
    wb.save(response)
    return response


def reporte_consumo_12_meses(request, alias):
    db_alias = f'db_{alias}'
    empresa = get_object_or_404(Empresa, slug=alias)

    user = request.user
    usuario_nombre = user.get_full_name() or user.username if user.is_authenticated else "Anónimo"
    santiago_tz = pytz.timezone('America/Santiago')
    ahora = datetime.now(santiago_tz)
    fecha_hora = ahora.strftime('%d/%m/%Y %H:%M:%S')
    fecha_archivo = ahora.strftime('%Y%m%d_%H%M%S')

    # Rango de fechas para el título (últimos 12 meses)
    hoy = ahora.date()
    fecha_inicio = hoy - timedelta(days=365)
    meses_es = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
    rango_titulo = f"{meses_es[fecha_inicio.month-1]} {fecha_inicio.year} - {meses_es[hoy.month-1]} {hoy.year}"

    lecturas = LecturaMovil.objects.filter(
        empresa_slug=alias,
        fecha_lectura__gte=fecha_inicio,
        consumo__isnull=False
    ).annotate(
        mes=TruncMonth('fecha_lectura')
    ).values('mes').annotate(
        total_consumo=Sum('consumo')
    ).order_by('mes')

    # Preparar datos para los últimos 12 meses (incluyendo ceros)
    meses_abbr = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    data = []
    for i in range(11, -1, -1):
        fecha = hoy - timedelta(days=30 * i)
        mes_num = fecha.month
        anio = fecha.year
        key = f"{anio}-{mes_num:02d}"
        total = next((item['total_consumo'] for item in lecturas if item['mes'].strftime('%Y-%m') == key), 0)
        data.append({'mes': f"{meses_abbr[mes_num-1]} {anio}", 'total': total})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consumo 12 Meses"
    ws.sheet_view.showGridLines = False

    titulo_reporte = "CONSUMO ÚLTIMOS 12 MESES"
    fila_inicio_tabla = escribir_cabecera_estandar(
        ws, empresa, usuario_nombre, fecha_hora, titulo_reporte, rango_titulo.upper()
    )

    # Estilos
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    data_font = Font(name='Calibri', size=9)
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_alignment_right = Alignment(horizontal='right', vertical='center')

    full_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # Encabezados
    headers = ['Mes', 'Consumo (m³)']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=fila_inicio_tabla, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = full_border

    # Datos
    for row_idx, item in enumerate(data, start=fila_inicio_tabla + 1):
        cell_mes = ws.cell(row=row_idx, column=1, value=item['mes'])
        cell_cons = ws.cell(row=row_idx, column=2, value=item['total'])
        for cell in (cell_mes, cell_cons):
            cell.font = data_font
            cell.alignment = data_alignment_right if isinstance(cell.value, (int, float)) else data_alignment
            cell.border = full_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill
        cell_cons.number_format = '#,##0.00'

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.freeze_panes = f'A{fila_inicio_tabla + 1}'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="consumo_12_meses_{alias}_{fecha_archivo}.xlsx"'
    wb.save(response)
    return response


def reporte_consumo_anual(request, alias):
    db_alias = f'db_{alias}'
    empresa = get_object_or_404(Empresa, slug=alias)

    user = request.user
    usuario_nombre = user.get_full_name() or user.username if user.is_authenticated else "Anónimo"
    santiago_tz = pytz.timezone('America/Santiago')
    ahora = datetime.now(santiago_tz)
    fecha_hora = ahora.strftime('%d/%m/%Y %H:%M:%S')
    fecha_archivo = ahora.strftime('%Y%m%d_%H%M%S')

    lecturas = LecturaMovil.objects.filter(
        empresa_slug=alias,
        consumo__isnull=False
    ).annotate(
        anio=TruncYear('fecha_lectura')
    ).values('anio').annotate(
        total_consumo=Sum('consumo')
    ).order_by('anio')

    # Rango de años disponibles para el subtítulo
    anios_disponibles = [item['anio'].year for item in lecturas if item['anio']]
    if anios_disponibles:
        rango_anios = f"{min(anios_disponibles)} - {max(anios_disponibles)}"
    else:
        rango_anios = "Sin datos"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consumo Anual"
    ws.sheet_view.showGridLines = False

    titulo_reporte = "RESUMEN CONSUMO ANUAL"
    fila_inicio_tabla = escribir_cabecera_estandar(
        ws, empresa, usuario_nombre, fecha_hora, titulo_reporte, rango_anios.upper()
    )

    # Estilos
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    data_font = Font(name='Calibri', size=9)
    data_alignment = Alignment(horizontal='center', vertical='center')
    data_alignment_right = Alignment(horizontal='right', vertical='center')

    full_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    headers = ['Año', 'Consumo Total (m³)']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=fila_inicio_tabla, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = full_border

    for row_idx, item in enumerate(lecturas, start=fila_inicio_tabla + 1):
        anio = item['anio'].year if item['anio'] else ''
        cell_anio = ws.cell(row=row_idx, column=1, value=anio)
        cell_cons = ws.cell(row=row_idx, column=2, value=item['total_consumo'])
        cell_anio.font = data_font
        cell_anio.alignment = data_alignment
        cell_anio.border = full_border
        cell_cons.font = data_font
        cell_cons.alignment = data_alignment_right
        cell_cons.border = full_border
        cell_cons.number_format = '#,##0.00'
        if row_idx % 2 == 0:
            cell_anio.fill = alt_fill
            cell_cons.fill = alt_fill

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.freeze_panes = f'A{fila_inicio_tabla + 1}'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="consumo_anual_{alias}_{fecha_archivo}.xlsx"'
    wb.save(response)
    return response

from clientes.models import Contrato

# lecturas/views.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from datetime import datetime
import pytz
from empresas.models import Empresa
from clientes.models import Cliente, Contrato
from lecturas.models import LecturaMovil

# Para PDF (opcional, solo si quieres usarlo)
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

def escribir_cabecera_estandar(ws, empresa, usuario_nombre, fecha_hora, titulo_reporte, subtitulo=None):
    """
    Escribe la cabecera estándar en la hoja Excel.
    Retorna el número de la siguiente fila disponible (donde empezará la tabla).
    """
    row = 1
    font_normal = Font(name='Calibri', size=11)
    font_bold = Font(name='Calibri', size=11, bold=True)
    font_titulo = Font(name='Calibri', size=12, bold=True)
    alignment_left = Alignment(horizontal='left', vertical='center')

    # Fila 1: Nombre comité
    ws.cell(row=row, column=1, value=empresa.nombre.upper()).font = font_bold
    ws.cell(row=row, column=1).alignment = alignment_left
    row += 1

    # Fila 2: RUT (si existe)
    rut_empresa = getattr(empresa, 'rut', '') or ''
    if rut_empresa:
        ws.cell(row=row, column=1, value=rut_empresa).font = font_normal
        ws.cell(row=row, column=1).alignment = alignment_left
        row += 1

    # Fila 3: Usuario
    ws.cell(row=row, column=1, value=f"Usuario: {usuario_nombre}").font = font_normal
    ws.cell(row=row, column=1).alignment = alignment_left
    row += 1

    # Fila 4: Fecha y hora
    ws.cell(row=row, column=1, value=fecha_hora).font = font_normal
    ws.cell(row=row, column=1).alignment = alignment_left
    row += 1

    # Fila 5: Título del reporte
    ws.cell(row=row, column=1, value=titulo_reporte.upper()).font = font_titulo
    ws.cell(row=row, column=1).alignment = alignment_left
    row += 1

    # Fila 6: Subtítulo (opcional)
    if subtitulo:
        ws.cell(row=row, column=1, value=subtitulo.upper()).font = font_normal
        ws.cell(row=row, column=1).alignment = alignment_left
        row += 1

    # Dejar una fila en blanco antes de la tabla
    row += 1

    return row

def reporte_lecturas_por_periodo(request, alias):
    db_alias = f'db_{alias}'
    empresa = get_object_or_404(Empresa, slug=alias)

    if request.method == 'POST':
        fecha_inicio_str = request.POST.get('fecha_inicio')
        fecha_fin_str = request.POST.get('fecha_fin')
        nombre_socio = request.POST.get('nombre_socio', '').strip()
        numero_contrato = request.POST.get('numero_contrato', '').strip()
        es_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.POST.get('ajax') == '1'

        if not (fecha_inicio_str and fecha_fin_str):
            if es_ajax:
                return JsonResponse({'success': False, 'error': 'Fechas requeridas'})
            return render(request, 'informes/reporte_periodo.html', {'empresa': empresa, 'slug': alias})

        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()

        user = request.user
        usuario_nombre = user.get_full_name() or user.username if user.is_authenticated else "Anónimo"
        santiago_tz = pytz.timezone('America/Santiago')
        ahora = datetime.now(santiago_tz)
        fecha_hora = ahora.strftime('%d/%m/%Y %H:%M:%S')
        fecha_archivo = ahora.strftime('%Y%m%d_%H%M%S')

        # Obtener lecturas del período
        lecturas = LecturaMovil.objects.filter(
            empresa_slug=alias,
            fecha_lectura__range=[fecha_inicio, fecha_fin],
            consumo__isnull=False
        ).order_by('fecha_lectura')

        # Clientes y contratos
        client_ids = list(set(lecturas.values_list('cliente', flat=True)))
        cliente_dict = {}
        contratos_dict = {}
        if client_ids:
            clientes = Cliente.objects.using(db_alias).filter(id__in=client_ids)
            cliente_dict = {c.id: c for c in clientes}
            contratos = Contrato.objects.using(db_alias).filter(cliente_id__in=client_ids)
            contratos_dict = {c.cliente_id: c for c in contratos}

        # Filtrar manualmente
        lecturas_filtradas = []
        for lectura in lecturas:
            cliente = cliente_dict.get(lectura.cliente)
            if not cliente:
                continue

            # Filtro por nombre de socio
            if nombre_socio:
                nombre_completo = f"{cliente.nombre} {cliente.apellido_paterno} {cliente.apellido_materno}".lower()
                if nombre_socio.lower() not in nombre_completo:
                    continue

            # Filtro por número de contrato (exacto)
            if numero_contrato:
                contrato = contratos_dict.get(cliente.id)
                if not contrato or not contrato.numero_contrato:
                    continue
                if contrato.numero_contrato.strip() != numero_contrato.strip():
                    continue

            lecturas_filtradas.append(lectura)

        # Si es AJAX, devolver JSON para el PDF
        if es_ajax:
            data_lecturas = []
            for lectura in lecturas_filtradas:
                cliente = cliente_dict.get(lectura.cliente)
                contrato = contratos_dict.get(lectura.cliente) if cliente else None
                data_lecturas.append({
                    'fecha': lectura.fecha_lectura.strftime('%d/%m/%Y'),
                    'cliente': cliente.nombre if cliente else '',
                    'rut': cliente.rut if cliente else '',
                    'contrato': contrato.numero_contrato if contrato else '',
                    'lectura_actual': float(lectura.lectura_actual) if lectura.lectura_actual else 0,
                    'consumo': float(lectura.consumo) if lectura.consumo else 0,
                    'estado': lectura.estado,
                    'usuario': lectura.usuario_app or '',
                })
            return JsonResponse({'success': True, 'lecturas': data_lecturas})

        # Si no es AJAX, generar Excel (comportamiento original)
        return generar_excel_lecturas(
            empresa, usuario_nombre, fecha_hora, fecha_archivo,
            fecha_inicio, fecha_fin, nombre_socio, numero_contrato,
            lecturas_filtradas, cliente_dict, contratos_dict
        )

    return render(request, 'informes/reporte_periodo.html', {'empresa': empresa, 'slug': alias})


def generar_excel_lecturas(empresa, usuario_nombre, fecha_hora, fecha_archivo,
                           fecha_inicio, fecha_fin, nombre_socio, numero_contrato,
                           lecturas_filtradas, cliente_dict, contratos_dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lecturas por Periodo"
    ws.sheet_view.showGridLines = False

    subtitulo = f"{fecha_inicio.strftime('%d/%m/%Y')} AL {fecha_fin.strftime('%d/%m/%Y')}"
    if nombre_socio:
        subtitulo += f" | Socio: {nombre_socio}"
    if numero_contrato:
        subtitulo += f" | Contrato: {numero_contrato}"

    titulo_reporte = "LECTURAS POR PERIODO"
    fila_inicio_tabla = escribir_cabecera_estandar(
        ws, empresa, usuario_nombre, fecha_hora, titulo_reporte, subtitulo.upper()
    )

    # Estilos
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='Calibri', size=9)
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_alignment_right = Alignment(horizontal='right', vertical='center')
    full_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    headers = ['Fecha', 'Cliente', 'RUT', 'N° Contrato', 'Lectura Actual (m³)', 'Consumo (m³)', 'Estado', 'Usuario']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=fila_inicio_tabla, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = full_border

    if lecturas_filtradas:
        for row_idx, lectura in enumerate(lecturas_filtradas, start=fila_inicio_tabla + 1):
            cliente = cliente_dict.get(lectura.cliente)
            contrato = contratos_dict.get(lectura.cliente) if cliente else None
            row_data = [
                lectura.fecha_lectura.strftime('%d/%m/%Y'),
                cliente.nombre if cliente else '',
                cliente.rut if cliente else '',
                contrato.numero_contrato if contrato else '',
                float(lectura.lectura_actual) if lectura.lectura_actual else '',
                float(lectura.consumo) if lectura.consumo else '',
                lectura.estado,
                lectura.usuario_app or '',
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.alignment = data_alignment_right if isinstance(val, (int, float)) else data_alignment
                cell.border = full_border
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    cell.number_format = '#,##0.00'
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
    else:
        ws.merge_cells(f'A{fila_inicio_tabla}:H{fila_inicio_tabla}')
        cell = ws.cell(fila_inicio_tabla, 1, 'No hay lecturas que coincidan con los filtros seleccionados.')
        cell.font = Font(size=12, bold=True, color='FF0000')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[fila_inicio_tabla].height = 30

    # Ajustar ancho de columnas
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        for row in range(fila_inicio_tabla, ws.max_row + 1):
            val = ws.cell(row, col_idx).value
            if val:
                max_length = max(max_length, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 3, 40)

    ws.freeze_panes = f'A{fila_inicio_tabla + 1}'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="lecturas_periodo_{empresa.slug}_{fecha_archivo}.xlsx"'
    wb.save(response)
    return response


def generar_pdf_lecturas(empresa, usuario_nombre, fecha_hora,
                         fecha_inicio, fecha_fin, nombre_socio, numero_contrato,
                         lecturas_filtradas, cliente_dict, contratos_dict):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="lecturas_periodo_{empresa.slug}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                            rightMargin=30, leftMargin=30,
                            topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    style_title = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=14, alignment=1, spaceAfter=10)
    style_subtitle = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=1, spaceAfter=20)
    style_table_header = ParagraphStyle('Header', parent=styles['Normal'], fontSize=9, alignment=1, textColor=colors.white)
    style_table_cell = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8)

    story = []
    story.append(Paragraph(f"LECTURAS POR PERIODO - {empresa.nombre.upper()}", style_title))
    subtitulo = f"{fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"
    if nombre_socio:
        subtitulo += f" | Socio: {nombre_socio}"
    if numero_contrato:
        subtitulo += f" | Contrato: {numero_contrato}"
    story.append(Paragraph(subtitulo, style_subtitle))
    story.append(Paragraph(f"Generado por: {usuario_nombre} | {fecha_hora}", styles['Normal']))
    story.append(Spacer(1, 0.2*inch))

    # Tabla
    headers = ['Fecha', 'Cliente', 'RUT', 'N° Contrato', 'Lectura Actual', 'Consumo', 'Estado', 'Usuario']
    data = [headers]
    for lectura in lecturas_filtradas:
        cliente = cliente_dict.get(lectura.cliente)
        contrato = contratos_dict.get(lectura.cliente) if cliente else None
        data.append([
            lectura.fecha_lectura.strftime('%d/%m/%Y'),
            cliente.nombre if cliente else '',
            cliente.rut if cliente else '',
            contrato.numero_contrato if contrato else '',
            f"{lectura.lectura_actual:.2f}" if lectura.lectura_actual else '',
            f"{lectura.consumo:.2f}" if lectura.consumo else '',
            lectura.estado,
            lectura.usuario_app or '',
        ])

    if not lecturas_filtradas:
        data.append(['', 'No hay datos para los filtros seleccionados', '', '', '', '', '', ''])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2E75B6')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F2F2F2')]),
    ]))
    story.append(table)

    doc.build(story)
    return response